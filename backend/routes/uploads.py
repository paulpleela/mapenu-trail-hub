"""
File upload routes
Handles GPX, LiDAR (.las/.laz), and XLSX file uploads
"""
from fastapi import APIRouter, File, UploadFile, HTTPException, Form
from typing import Optional
import gpxpy
import os
import uuid
import tempfile
from database import supabase, supabase_service
from utils.calculations import haversine, analyze_rolling_hills
from utils.terrain_analysis import get_trail_weather_exposure, calculate_terrain_variety

router = APIRouter()


@router.post("/upload-gpx")
async def upload_gpx(file: UploadFile = File(...), overwrite: str = Form("false")):
    """Handle GPX file upload and save to Supabase

    Args:
        file: GPX file to upload
        overwrite: If True, will delete existing trail with same name/location before adding new one
    """
    # Convert overwrite string to boolean
    print(f"📤 Uploading GPX file: {file.filename}")
    print(f"   Raw overwrite value: '{overwrite}' (type: {type(overwrite).__name__})")

    # Safety check for None or empty string
    if overwrite is None:
        overwrite_bool = False
    elif isinstance(overwrite, str):
        overwrite_bool = overwrite.lower() in ("true", "1", "yes")
    else:
        overwrite_bool = bool(overwrite)

    print(f"   Converted to bool: {overwrite_bool}")

    if not file.filename.lower().endswith(".gpx"):
        raise HTTPException(status_code=400, detail="File must be a GPX file")

    try:
        # Read GPX content
        content = await file.read()
        gpx_content = content.decode("utf-8")

        # Extract trail name from filename
        trail_name = file.filename.replace(".gpx", "").replace("_", " ").title()

        # Analyze trail data
        gpx = gpxpy.parse(gpx_content)
        coords = []
        elevations = []
        distances = [0]
        slopes = [0]  # Start with 0 slope for first point

        for track in gpx.tracks:
            for segment in track.segments:
                for i, point in enumerate(segment.points):
                    coords.append([point.latitude, point.longitude])
                    elevations.append(point.elevation or 0)

                    if i > 0:
                        prev_point = segment.points[i - 1]
                        dist = (
                            haversine(
                                prev_point.latitude,
                                prev_point.longitude,
                                point.latitude,
                                point.longitude,
                            )
                            / 1000
                        )
                        distances.append(distances[-1] + dist)

                        # Slope analysis (gradient in %)
                        elev_diff = (point.elevation or 0) - (prev_point.elevation or 0)
                        dist_m = haversine(
                            prev_point.latitude,
                            prev_point.longitude,
                            point.latitude,
                            point.longitude,
                        )
                        if dist_m > 0:
                            gradient = (elev_diff / dist_m) * 100
                            slopes.append(gradient)
                        else:
                            slopes.append(0)

        if not coords:
            raise HTTPException(
                status_code=400, detail="No track points found in GPX file"
            )

        # Calculate statistics
        total_distance = distances[-1] if len(distances) > 1 else 0
        elevation_gain = (
            sum(
                max(0, elevations[i] - elevations[i - 1])
                for i in range(1, len(elevations))
            )
            if len(elevations) > 1
            else 0
        )
        elevation_loss = (
            sum(
                max(0, elevations[i - 1] - elevations[i])
                for i in range(1, len(elevations))
            )
            if len(elevations) > 1
            else 0
        )
        max_elevation = max(elevations) if elevations else 0
        min_elevation = min(elevations) if elevations else 0

        # Rolling hills analysis (advanced) - returns index and count
        rolling_hills_index, rolling_hills_count = analyze_rolling_hills(
            elevations, distances
        )
        rolling_hills_index = round(rolling_hills_index, 2)
        print(f"🔍 DEBUG: Rolling Hills Index calculated: {rolling_hills_index}")
        print(f"🔍 DEBUG: Rolling Hills Count: {rolling_hills_count}")
        print(
            f"🔍 DEBUG: Elevations count: {len(elevations)}, Distance: {distances[-1] if distances else 0} km"
        )

        # Create elevation profile data
        elevation_profile_data = [
            {
                "distance": round(dist, 2),
                "elevation": round(ele, 1),
                "slope": round(slopes[i] if i < len(slopes) else 0, 2),
            }
            for i, (dist, ele) in enumerate(zip(distances, elevations))
        ]

        # Slope analysis
        if slopes and len(slopes) > 1:  # Skip first element (always 0)
            slope_values = slopes[1:]  # Skip the initial 0
            max_slope = max(slope_values) if slope_values else 0
            avg_slope = (
                sum(map(abs, slope_values)) / len(slope_values) if slope_values else 0
            )
        else:
            max_slope = 0
            avg_slope = 0

        # Segment analysis (500m segments)
        segment_length = 0.5  # km
        segments = []
        if len(distances) > 1:
            seg_start_idx = 0
            while seg_start_idx < len(distances) - 1:
                seg_end_idx = seg_start_idx
                # Find the end index for this segment
                while (
                    seg_end_idx < len(distances) - 1
                    and distances[seg_end_idx] - distances[seg_start_idx]
                    < segment_length
                ):
                    seg_end_idx += 1
                # Calculate stats for this segment
                seg_dist = distances[seg_end_idx] - distances[seg_start_idx]
                seg_elev_change = elevations[seg_end_idx] - elevations[seg_start_idx]
                # Slope for segment
                if seg_dist > 0:
                    seg_slope = (seg_elev_change / (seg_dist * 1000)) * 100
                else:
                    seg_slope = 0
                segments.append(
                    {
                        "start_distance": round(distances[seg_start_idx], 2),
                        "end_distance": round(distances[seg_end_idx], 2),
                        "elevation_change": round(seg_elev_change, 1),
                        "avg_slope": round(seg_slope, 2),
                    }
                )
                seg_start_idx = seg_end_idx

        # Simple difficulty calculation
        distance_factor = min(total_distance / 10, 1) * 3  # 0-3 points
        elevation_factor = min(elevation_gain / 1000, 1) * 4  # 0-4 points
        # Normalize rolling hills index (typically 0-50 range) to 0-1, then scale to 0-3 points
        normalized_rolling = min(
            rolling_hills_index / 50, 1
        )  # Cap at 50 for normalization
        rolling_factor = normalized_rolling * 3  # 0-3 points
        difficulty_score = (
            distance_factor + elevation_factor + rolling_factor
        )  # Max = 10 points

        if difficulty_score <= 3:
            difficulty_level = "Easy"
        elif difficulty_score <= 6:
            difficulty_level = "Moderate"
        elif difficulty_score <= 8:
            difficulty_level = "Hard"
        else:
            difficulty_level = "Extreme"

        # Check for duplicate trails before inserting
        # First check by exact name match
        existing_trails_response = (
            supabase.table("trails").select("*").eq("name", trail_name).execute()
        )

        duplicate_trail_id = None
        if existing_trails_response.data:
            if not overwrite_bool:
                raise HTTPException(
                    status_code=409,
                    detail=f"Trail with name '{trail_name}' already exists in database",
                )
            else:
                # Delete existing trail and its associated LiDAR files
                duplicate_trail_id = existing_trails_response.data[0]["id"]
                print(
                    f"🗑️  Overwrite mode: Deleting existing trail ID {duplicate_trail_id}"
                )

                # Delete associated LiDAR files first
                lidar_files = (
                    supabase.table("lidar_files")
                    .select("*")
                    .eq("trail_id", duplicate_trail_id)
                    .execute()
                )
                if lidar_files.data:
                    print(
                        f"   Deleting {len(lidar_files.data)} associated LiDAR file(s)"
                    )
                    for lidar_file in lidar_files.data:
                        db_client = supabase_service if supabase_service else supabase
                        db_client.table("lidar_files").delete().eq(
                            "id", lidar_file["id"]
                        ).execute()

                # Delete the trail
                db_client = supabase_service if supabase_service else supabase
                db_client.table("trails").delete().eq(
                    "id", duplicate_trail_id
                ).execute()
                print(f"   ✅ Deleted trail and associated data")

        # Check for similar starting coordinates (within ~100m radius)
        # This prevents uploading the same trail with different names
        start_lat, start_lon = coords[0]
        all_trails_response = (
            supabase.table("trails").select("id, name, coordinates").execute()
        )

        for existing_trail in all_trails_response.data:
            if existing_trail.get("coordinates"):
                existing_start = existing_trail["coordinates"][0]
                existing_lat, existing_lon = existing_start

                # Calculate distance between starting points
                distance_between_starts = haversine(
                    start_lat, start_lon, existing_lat, existing_lon
                )

                # If starts are within 100 meters, likely duplicate
                if distance_between_starts < 100:
                    if not overwrite_bool:
                        raise HTTPException(
                            status_code=409,
                            detail=f"Trail starting near same location as existing trail '{existing_trail['name']}' (within 100m). Possible duplicate.",
                        )
                    else:
                        # Delete this coordinate-duplicate trail too
                        coord_dup_id = existing_trail["id"]
                        if coord_dup_id != duplicate_trail_id:  # Don't delete twice
                            print(
                                f"🗑️  Overwrite mode: Deleting coordinate-duplicate trail ID {coord_dup_id}"
                            )

                            # Delete associated LiDAR files
                            lidar_files = (
                                supabase.table("lidar_files")
                                .select("*")
                                .eq("trail_id", coord_dup_id)
                                .execute()
                            )
                            if lidar_files.data:
                                print(
                                    f"   Deleting {len(lidar_files.data)} associated LiDAR file(s)"
                                )
                                for lidar_file in lidar_files.data:
                                    db_client = (
                                        supabase_service
                                        if supabase_service
                                        else supabase
                                    )
                                    db_client.table("lidar_files").delete().eq(
                                        "id", lidar_file["id"]
                                    ).execute()

                            # Delete the trail
                            db_client = (
                                supabase_service if supabase_service else supabase
                            )
                            db_client.table("trails").delete().eq(
                                "id", coord_dup_id
                            ).execute()
                            print(f"   ✅ Deleted coordinate-duplicate trail")

        # Create new trail data for Supabase
        weather_exposure = get_trail_weather_exposure({"max_elevation": max_elevation})
        terrain_variety = calculate_terrain_variety(elevations)

        # Convert weather exposure to a numeric score for database compatibility
        exposure_scores = {
            "Low": 1.0,
            "Low-Moderate": 1.1,
            "Moderate": 1.2,
            "High": 1.3,
        }
        weather_score = exposure_scores.get(weather_exposure["exposure_level"], 1.0)

        new_trail_data = {
            "name": trail_name,
            "distance": round(total_distance, 2),
            "elevation_gain": int(round(elevation_gain, 0)),
            "elevation_loss": int(round(elevation_loss, 0)),
            "max_elevation": int(round(max_elevation, 0)),
            "min_elevation": int(round(min_elevation, 0)),
            "rolling_hills_index": rolling_hills_index,
            "rolling_hills_count": rolling_hills_count,  # Number of significant elevation changes
            "difficulty_score": round(difficulty_score, 1),
            "difficulty_level": difficulty_level,
            "coordinates": coords,
            "elevation_profile": elevation_profile_data,
            "max_slope": round(max_slope, 2),
            "avg_slope": round(avg_slope, 2),
            "segments": segments,
            # Enhanced effort estimation using Naismith's Rule + terrain adjustments
            "estimated_time_hours": round(
                (total_distance / 5)
                + (elevation_gain / 600)
                + (rolling_hills_index * 0.5),
                2,
            ),
            # Improved analytics fields (using existing column names)
            "terrain_variety_score": terrain_variety,
            "elevation_change_total": int(round(elevation_gain + elevation_loss, 0)),
            # Store weather data in existing numeric fields, we'll interpret on frontend
            "weather_difficulty_multiplier": weather_score,  # Use existing column with new meaning
            # Fixed technical difficulty calculation (1-10 scale)
            # Factors: max slope (40%), rolling hills (30%), avg slope (30%)
            # Properly normalized to produce 1-10 range
            "technical_rating": round(
                max(
                    1.0,
                    min(
                        10.0,
                        1
                        + (max_slope / 100) * 3.5  # Max slope: 0-100% -> 0-3.5 points
                        + (
                            min(rolling_hills_index / 50, 1.0) * 3.5
                        )  # Rolling hills normalized: 0-50 -> 0-3.5 points
                        + (avg_slope / 30) * 2.0,  # Avg slope: 0-30% -> 0-2 points
                    ),
                ),
            ),
        }

        # Insert trail into Supabase database (use service-role client if available to bypass RLS)
        db_client = supabase_service if supabase_service else supabase
        response = db_client.table("trails").insert(new_trail_data).execute()

        if response.data:
            inserted_trail = response.data[0]
            return {
                "success": True,
                "message": "Trail uploaded successfully to database",
                "trail": inserted_trail,
            }
        else:
            raise HTTPException(
                status_code=500, detail="Failed to insert trail into database"
            )

    except HTTPException:
        # Re-raise HTTPExceptions (409, 400, etc.) without wrapping them
        raise
    except Exception as e:
        import traceback

        error_details = traceback.format_exc()
        print(f"❌ Upload error: {type(e).__name__}: {e}")
        print(f"❌ Full traceback:\n{error_details}")
        raise HTTPException(
            status_code=500, detail=f"{type(e).__name__}: {str(e) or 'Unknown error'}"
        )


@router.post("/upload-lidar")
async def upload_lidar_file(
    file: UploadFile = File(...),
    trail_id: Optional[int] = Form(None),
    overwrite: str = Form("false"),  # Receive as string, convert to bool below
):
    """
    Upload a LiDAR .las/.laz file to Supabase Storage and store metadata in the database
    Checks for duplicates and file size before uploading

    Args:
        file: LiDAR .las or .laz file (compressed or uncompressed)
        trail_id: Optional trail ID to associate with this LiDAR file
        overwrite: If True, will delete existing LiDAR for this trail before uploading
    """
    temp_path = None
    try:
        # Convert overwrite string to boolean
        print(f"📤 Uploading LiDAR file: {file.filename}")
        print(f"   Trail ID: {trail_id}")
        print(
            f"   Raw overwrite value: '{overwrite}' (type: {type(overwrite).__name__})"
        )

        # Safety check for None or empty string
        if overwrite is None:
            overwrite_bool = False
        elif isinstance(overwrite, str):
            overwrite_bool = overwrite.lower() in ("true", "1", "yes")
        else:
            overwrite_bool = bool(overwrite)

        print(f"   Converted to bool: {overwrite_bool}")

        # Validate file extension
        if not (file.filename.endswith(".las") or file.filename.endswith(".laz")):
            raise HTTPException(
                status_code=400,
                detail="Invalid file format. Only .las and .laz files are supported.",
            )

        # Read file content to check size
        print(f"📊 Reading file to check size...")
        content = await file.read()
        file_size_mb = len(content) / (1024 * 1024)

        # Check if file is too large (>= 1GB)
        if file_size_mb >= 1024:
            raise HTTPException(
                status_code=413,
                detail=f"File too large ({file_size_mb:.0f} MB). Files >= 1GB must be added using the add_local_lidar_to_db.py script. "
                f"Run: python3 add_local_lidar_to_db.py 'path/to/{file.filename}' {trail_id or '<trail_id>'}",
            )

        # Check if trail already has LiDAR file(s)
        print(
            f"🔍 Checking trail LiDAR: trail_id={trail_id}, overwrite_bool={overwrite_bool}"
        )
        if trail_id and not overwrite_bool:
            existing_trail_lidar = (
                supabase.table("lidar_files")
                .select("id, filename")
                .eq("trail_id", trail_id)
                .execute()
            )
            print(
                f"🔍 Existing trail LiDAR check: {len(existing_trail_lidar.data) if existing_trail_lidar.data else 0} file(s) found"
            )
            if existing_trail_lidar.data:
                print(f"⚠️  Trail already has LiDAR - raising 409 error")
                raise HTTPException(
                    status_code=409,
                    detail=f"Trail ID {trail_id} already has LiDAR file(s): {', '.join([f['filename'] for f in existing_trail_lidar.data])}. "
                    "Use overwrite=true to replace existing files.",
                )

        # Check for duplicates in database (by original filename)
        print(f"🔍 Checking for existing file: {file.filename}")
        existing_check = (
            supabase.table("lidar_files")
            .select("id, filename, file_url, trail_id")
            .ilike("filename", f"%{file.filename}")
            .execute()
        )

        if existing_check.data:
            if not overwrite_bool:
                existing_file = existing_check.data[0]
                print(f"⚠️  File already exists: {existing_file['filename']}")
                raise HTTPException(
                    status_code=409,
                    detail=f"File '{file.filename}' already exists in database. "
                    f"Existing file: {existing_file['filename']}. "
                    f"Please rename your file or delete the existing one first.",
                )
            else:
                # Delete existing file(s) with same name
                for existing_file in existing_check.data:
                    print(
                        f"🗑️  Overwrite mode: Deleting existing file ID {existing_file['id']}"
                    )
                    try:
                        # Delete from storage
                        if existing_file["file_url"] and not existing_file[
                            "file_url"
                        ].startswith("local://"):
                            supabase.storage.from_("lidar-files").remove(
                                [existing_file["filename"]]
                            )
                    except Exception as e:
                        print(f"   ⚠️  Could not delete from storage: {e}")
                    # Delete from database
                    db_client = supabase_service if supabase_service else supabase
                    db_client.table("lidar_files").delete().eq(
                        "id", existing_file["id"]
                    ).execute()
                    print(f"   ✅ Deleted existing file")

        # If overwriting trail's LiDAR, delete existing trail LiDAR files
        if trail_id and overwrite_bool:
            existing_trail_lidar = (
                supabase.table("lidar_files")
                .select("id, filename, file_url")
                .eq("trail_id", trail_id)
                .execute()
            )
            if existing_trail_lidar.data:
                print(
                    f"🗑️  Overwrite mode: Deleting {len(existing_trail_lidar.data)} existing LiDAR file(s) for trail {trail_id}"
                )
                for lidar_file in existing_trail_lidar.data:
                    try:
                        # Delete from storage
                        if lidar_file["file_url"] and not lidar_file[
                            "file_url"
                        ].startswith("local://"):
                            supabase.storage.from_("lidar-files").remove(
                                [lidar_file["filename"]]
                            )
                    except Exception as e:
                        print(f"   ⚠️  Could not delete from storage: {e}")
                    # Delete from database
                    db_client = supabase_service if supabase_service else supabase
                    db_client.table("lidar_files").delete().eq(
                        "id", lidar_file["id"]
                    ).execute()
                print(f"   ✅ Deleted existing trail LiDAR files")

        # Generate unique filename to avoid conflicts in storage
        timestamp = uuid.uuid4().hex[:8]
        safe_filename = f"{timestamp}_{file.filename}"

        print(f"📊 File size: {file_size_mb:.2f} MB")

        # Upload to Supabase Storage
        print(f"☁️  Uploading {safe_filename} to Supabase Storage...")
        try:
            storage_response = supabase.storage.from_("lidar-files").upload(
                path=safe_filename,
                file=content,
                file_options={"content-type": "application/octet-stream"},
            )
            print(f"✅ Upload response: {storage_response}")
        except Exception as storage_error:
            print(f"❌ Storage upload error: {storage_error}")
            raise HTTPException(
                status_code=500,
                detail=f"Failed to upload to storage: {str(storage_error)}",
            )

        # Get public URL
        file_url = supabase.storage.from_("lidar-files").get_public_url(safe_filename)
        print(f"🔗 File URL: {file_url}")

        # Save temporarily to extract metadata
        temp_path = os.path.join("/tmp", safe_filename)
        print(f"💾 Saving temporary file for metadata extraction: {temp_path}")
        with open(temp_path, "wb") as f:
            f.write(content)

        # Extract metadata from LiDAR file
        try:
            import laspy

            las_data = laspy.open(temp_path)
            header = las_data.header

            metadata = {
                "filename": safe_filename,
                "original_filename": file.filename,
                "file_url": file_url,
                "file_size_mb": round(file_size_mb, 2),
                "point_count": header.point_count,
                "min_x": float(header.x_min),
                "max_x": float(header.x_max),
                "min_y": float(header.y_min),
                "max_y": float(header.y_max),
                "min_z": float(header.z_min),
                "max_z": float(header.z_max),
                "las_version": f"{header.version.major}.{header.version.minor}",
                "point_format_id": header.point_format.id,
                "crs_epsg": 28356,  # Assuming GDA94 MGA Zone 56
            }

            las_data.close()
            print(f"📈 Metadata extracted: {header.point_count} points")

        except Exception as e:
            print(f"⚠️  Error extracting LiDAR metadata: {e}")
            # Use basic metadata if extraction fails
            metadata = {
                "filename": safe_filename,
                "original_filename": file.filename,
                "file_url": file_url,
                "file_size_mb": round(file_size_mb, 2),
                "point_count": None,
                "min_x": None,
                "max_x": None,
                "min_y": None,
                "max_y": None,
                "min_z": None,
                "max_z": None,
                "las_version": None,
                "point_format_id": None,
                "crs_epsg": 28356,
            }

        # Store metadata in Supabase database
        lidar_record = {
            "trail_id": trail_id,
            "filename": metadata["filename"],
            "file_url": metadata["file_url"],  # Store URL instead of path
            "file_size_mb": metadata["file_size_mb"],
            "point_count": metadata.get("point_count"),
            "min_x": metadata.get("min_x"),
            "max_x": metadata.get("max_x"),
            "min_y": metadata.get("min_y"),
            "max_y": metadata.get("max_y"),
            "min_z": metadata.get("min_z"),
            "max_z": metadata.get("max_z"),
            "las_version": metadata.get("las_version"),
            "point_format_id": metadata.get("point_format_id"),
            "crs_epsg": metadata.get("crs_epsg"),
        }

        try:
            db_response = supabase.table("lidar_files").insert(lidar_record).execute()
            print(f"💾 LiDAR file metadata saved to database: {db_response.data}")
        except Exception as db_error:
            print(f"❌ Database error: {db_error}")
            raise HTTPException(
                status_code=500,
                detail=f"Failed to save metadata to database: {str(db_error)}",
            )

        # Reinitialize LiDAR extractor to pick up new file
        import app_state

        lidar_extractor = app_state.get_lidar_extractor()
        if lidar_extractor:
            lidar_extractor.lidar_files = lidar_extractor._find_lidar_files()
            print(
                f"🔄 LiDAR extractor reinitialized with {len(lidar_extractor.lidar_files)} files"
            )

        return {
            "success": True,
            "message": "LiDAR file uploaded successfully to Supabase Storage",
            "file_url": file_url,
            "metadata": metadata,
            "database_record": db_response.data[0] if db_response.data else None,
        }

    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ LiDAR upload error: {e}")
        import traceback

        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        # Clean up temporary file
        if temp_path and os.path.exists(temp_path):
            try:
                os.remove(temp_path)
                print(f"🗑️  Cleaned up temporary file: {temp_path}")
            except Exception as cleanup_error:
                print(f"⚠️  Could not clean up temp file: {cleanup_error}")


@router.post("/upload-xlsx")
async def upload_xlsx_file(
    file: UploadFile = File(...),
    trail_id: Optional[int] = Form(None),
    overwrite: str = Form("false"),
):
    """
    Upload an XLSX file to Supabase Storage and store metadata in the database
    Expects a spreadsheet with columns including 'layer', 'distance', 'elevation'
    """
    temp_path = None
    try:
        print(f"📤 Uploading XLSX file: {file.filename}")
        # Convert overwrite to bool
        if overwrite is None:
            overwrite_bool = False
        elif isinstance(overwrite, str):
            overwrite_bool = overwrite.lower() in ("true", "1", "yes")
        else:
            overwrite_bool = bool(overwrite)

        # Validate extension
        if not (file.filename.endswith(".xlsx") or file.filename.endswith(".xls")):
            raise HTTPException(
                status_code=400, detail="Invalid file type. Use .xlsx or .xls"
            )

        content = await file.read()
        file_size_mb = len(content) / (1024 * 1024)

        # Upload to Supabase Storage
        timestamp = uuid.uuid4().hex[:8]
        safe_filename = f"{timestamp}_{file.filename}"
        try:
            storage_response = supabase.storage.from_("xlsx-files").upload(
                path=safe_filename,
                file=content,
                file_options={
                    "content-type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                },
            )
            print(f"✅ XLSX upload response: {storage_response}")
        except Exception as e:
            print(f"❌ XLSX storage upload failed: {e}")
            raise HTTPException(status_code=500, detail=str(e))

        file_url = supabase.storage.from_("xlsx-files").get_public_url(safe_filename)

        # Save temporarily to inspect
        temp_path = os.path.join(tempfile.gettempdir(), safe_filename)
        with open(temp_path, "wb") as f:
            f.write(content)

        # Read sheet info and number of rows
        try:
            from openpyxl import load_workbook

            wb = load_workbook(filename=temp_path, read_only=True, data_only=True)
            sheet_name = wb.sheetnames[0] if wb.sheetnames else ""
            ws = wb[sheet_name] if sheet_name else None
            num_rows = 0
            if ws:
                for _ in ws.rows:
                    num_rows += 1
            else:
                num_rows = 0
            wb.close()
        except Exception as e:
            print(f"⚠️  Failed to read XLSX with openpyxl: {e}")
            sheet_name = ""
            num_rows = None

        # Insert metadata into database
        xlsx_record = {
            "trail_id": trail_id,
            "filename": safe_filename,
            "original_filename": file.filename,
            "file_url": file_url,
            "file_size_mb": round(file_size_mb, 2),
            "num_rows": num_rows,
            "sheet_name": sheet_name,
        }

        try:
            db_client = supabase_service if supabase_service else supabase
            db_resp = db_client.table("xlsx_files").insert(xlsx_record).execute()
            print(f"💾 XLSX metadata saved: {db_resp.data}")
        except Exception as e:
            print(f"❌ Failed to save XLSX metadata: {e}")
            raise HTTPException(status_code=500, detail=str(e))

        return {
            "success": True,
            "file_url": file_url,
            "metadata": xlsx_record,
            "db_record": db_resp.data,
        }

    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ XLSX upload error: {e}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if temp_path and os.path.exists(temp_path):
            try:
                os.remove(temp_path)
            except Exception:
                pass


@router.get("/lidar-files")
async def get_lidar_files():
    """Get list of all LiDAR files from database"""
    try:
        response = supabase.table("lidar_files").select("*").execute()
        return {
            "success": True,
            "lidar_files": response.data,
            "count": len(response.data),
        }
    except Exception as e:
        print(f"Error fetching LiDAR files: {e}")
        return {"success": False, "error": str(e), "lidar_files": [], "count": 0}


@router.delete("/lidar-files/{lidar_id}")
async def delete_lidar_file(lidar_id: int):
    """
    Delete a LiDAR file from both Supabase Storage and database

    Args:
        lidar_id: ID of the LiDAR file record to delete
    """
    try:
        # Get file info from database
        print(f"🔍 Looking up LiDAR file with ID: {lidar_id}")
        file_response = (
            supabase.table("lidar_files").select("*").eq("id", lidar_id).execute()
        )

        if not file_response.data:
            raise HTTPException(
                status_code=404, detail=f"LiDAR file with ID {lidar_id} not found"
            )

        file_record = file_response.data[0]
        filename = file_record.get("filename")
        file_url = file_record.get("file_url")

        print(f"📂 Found file: {filename}")

        # Delete from storage (if not local file)
        if file_url and not file_url.startswith("local://") and filename:
            try:
                supabase.storage.from_("lidar-files").remove([filename])
                print(f"✅ Deleted from storage: {filename}")
            except Exception as e:
                print(f"⚠️  Could not delete from storage: {filename} - {e}")

        # Delete from database
        db_client = supabase_service if supabase_service else supabase
        db_client.table("lidar_files").delete().eq("id", lidar_id).execute()
        print(f"✅ Deleted from database")

        # Reinitialize LiDAR extractor
        import app_state

        lidar_extractor = app_state.get_lidar_extractor()
        if lidar_extractor:
            lidar_extractor.lidar_files = lidar_extractor._find_lidar_files()
            print(f"🔄 LiDAR extractor reinitialized")

        return {
            "success": True,
            "message": f"LiDAR file '{filename}' deleted successfully",
            "deleted_file": file_record,
        }

    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Error deleting LiDAR file: {e}")
        import traceback

        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))
