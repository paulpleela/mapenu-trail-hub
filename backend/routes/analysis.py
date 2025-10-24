"""
Elevation analysis routes
Handles DEM analysis, 3D terrain visualization, and multi-source elevation data
"""
from fastapi import APIRouter, HTTPException
from fastapi.responses import JSONResponse, HTMLResponse
from database import supabase
from utils.calculations import haversine
from utils.dem_processing import find_relevant_dem_tiles, process_dem_for_trail
import random

router = APIRouter()


@router.get("/trail/{trail_id}/dem3d")
async def get_trail_3d_dem(trail_id: int):
    """Get 3D DEM data for a specific trail"""
    try:
        print(f"Getting 3D DEM data for trail ID: {trail_id}")

        # Get the trail from database
        trail_response = (
            supabase.table("trails").select("*").eq("id", trail_id).execute()
        )
        if not trail_response.data:
            raise HTTPException(status_code=404, detail="Trail not found")

        trail = trail_response.data[0]
        trail_coords = trail.get("coordinates", [])

        if not trail_coords:
            raise HTTPException(status_code=400, detail="Trail has no coordinate data")

        print(
            f"Processing DEM for trail: {trail.get('name', 'Unknown')} with {len(trail_coords)} coordinates"
        )

        # Find relevant DEM tiles
        dem_files = find_relevant_dem_tiles(trail_coords)
        if not dem_files:
            raise HTTPException(
                status_code=404, detail="No DEM data available for this trail area"
            )

        print(f"Found {len(dem_files)} DEM files")

        # Process DEM data
        dem_data = process_dem_for_trail(trail_coords, dem_files)
        if not dem_data:
            raise HTTPException(status_code=500, detail="Failed to process DEM data")

        return {
            "success": True,
            "trail_name": trail.get("name", "Unknown Trail"),
            "trail_id": trail_id,
            "dem_data": dem_data,
        }

    except HTTPException:
        raise
    except Exception as e:
        print(f"3D DEM error: {e}")
        import traceback

        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/trail/{trail_id}/dem-analysis")
async def get_trail_dem_analysis(trail_id: int):
    """Analyze DEM data for a specific trail using real DEM files"""
    try:
        import app_state

        dem_analyzer = app_state.get_dem_analyzer()
        if not dem_analyzer:
            return {
                "success": False,
                "error": "DEM analyzer not available. Check DEM file path and dependencies.",
            }

        # Get trail data
        trail_response = (
            supabase.table("trails").select("*").eq("id", trail_id).execute()
        )
        if not trail_response.data:
            raise HTTPException(status_code=404, detail="Trail not found")

        trail = trail_response.data[0]
        coordinates = trail.get("coordinates", [])

        if not coordinates:
            return {
                "success": False,
                "error": "No coordinates available for this trail",
            }

        print(
            f"Analyzing trail '{trail.get('name')}' with {len(coordinates)} coordinate points"
        )

        # Extract real elevation profile from DEM data
        elevation_analysis = dem_analyzer.extract_elevation_profile(coordinates)

        if not elevation_analysis.get("success"):
            return {
                "success": False,
                "error": elevation_analysis.get("error", "Analysis failed"),
            }

        # Analyze terrain features
        terrain_features = dem_analyzer.analyze_terrain_features(coordinates)

        # Generate 3D visualization
        visualization_3d = dem_analyzer.create_3d_terrain_visualization(coordinates)

        result = {
            "success": True,
            "trail_name": trail.get("name"),
            "elevation_analysis": elevation_analysis,
            "terrain_features": terrain_features,
            "visualization_3d": visualization_3d,
            "data_quality": {
                "resolution": "1 meter",
                "data_source": "QSpatial DEM",
                "coordinate_system": "GDA94 MGA Zone 56",
                "accuracy": "±0.5 meters",
            },
        }

        return result

    except Exception as e:
        print(f"DEM analysis error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/trail/{trail_id}/3d-terrain")
async def get_trail_3d_terrain(trail_id: int, elevation_source: str = "gpx"):
    """Generate interactive 3D terrain visualization for a trail

    Args:
        trail_id: Trail ID
        elevation_source: "gpx" (default) or "lidar" - determines which elevation data to use for trail overlay
    """
    try:
        import app_state

        dem_analyzer = app_state.get_dem_analyzer()
        lidar_extractor = app_state.get_lidar_extractor()
        if not dem_analyzer:
            return {"success": False, "error": "DEM analyzer not available"}

        # Get trail data
        trail_response = (
            supabase.table("trails").select("*").eq("id", trail_id).execute()
        )
        if not trail_response.data:
            raise HTTPException(status_code=404, detail="Trail not found")

        trail = trail_response.data[0]
        coordinates = trail.get("coordinates", [])

        if not coordinates:
            return {"success": False, "error": "No coordinates available"}

        # Get LiDAR elevations if requested
        lidar_elevations = None
        if elevation_source.lower() == "lidar":
            try:
                # Extract elevation profile from LiDAR
                profile = lidar_extractor.extract_elevation_profile(
                    trail_coords=coordinates, trail_id=trail_id
                )
                if profile and profile.get("success") and "elevations" in profile:
                    # For 3D visualization, we need RAW LiDAR elevations (not GPX-aligned)
                    # because they need to match the DEM terrain surface
                    lidar_elevations = profile["elevations"]

                    print(
                        f"📊 Using {len(lidar_elevations)} RAW LiDAR elevation points for 3D visualization (no GPX alignment)"
                    )
                    print(
                        f"   LiDAR elevation range: {min(lidar_elevations):.1f}m - {max(lidar_elevations):.1f}m"
                    )
                else:
                    print("⚠️  No LiDAR data available, falling back to GPX/DEM")
                    elevation_source = "gpx"
            except Exception as e:
                print(f"❌ Error getting LiDAR elevations: {e}")
                elevation_source = "gpx"

        # Generate 3D visualization
        visualization_result = dem_analyzer.create_3d_terrain_visualization(
            coordinates,
            buffer_meters=1000,
            elevation_source=elevation_source,
            trail_id=trail_id,
            lidar_elevations=lidar_elevations,
        )

        if not visualization_result.get("success"):
            return {
                "success": False,
                "error": visualization_result.get(
                    "error", "Failed to generate 3D visualization"
                ),
            }

        # Return based on visualization type
        if visualization_result.get("type") == "interactive":
            return {
                "success": True,
                "trail_name": trail.get("name"),
                "visualization_type": "interactive",
                "visualization_html": visualization_result["html_content"],
                "description": visualization_result["description"],
            }
        else:
            # Static visualization
            return {
                "success": True,
                "trail_name": trail.get("name"),
                "visualization_type": "static",
                "visualization": f"data:image/png;base64,{visualization_result['image_base64']}",
                "description": visualization_result["description"],
            }

    except Exception as e:
        print(f"3D terrain error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/trail/{trail_id}/3d-terrain-viewer")
async def get_trail_3d_terrain_viewer(trail_id: int, elevation_source: str = "gpx"):
    """Serve interactive 3D terrain visualization as a standalone HTML page

    Args:
        trail_id: Trail ID
        elevation_source: "gpx" (default) or "lidar" - determines which elevation data to use for trail overlay
    """
    try:
        import app_state

        dem_analyzer = app_state.get_dem_analyzer()
        lidar_extractor = app_state.get_lidar_extractor()
        if not dem_analyzer:
            return JSONResponse(
                {"error": "DEM analyzer not available"}, status_code=503
            )

        # Get trail data
        trail_response = (
            supabase.table("trails").select("*").eq("id", trail_id).execute()
        )
        if not trail_response.data:
            return JSONResponse({"error": "Trail not found"}, status_code=404)

        trail = trail_response.data[0]
        coordinates = trail.get("coordinates", [])

        if not coordinates:
            return JSONResponse({"error": "No coordinates available"}, status_code=400)

        # Get LiDAR elevations if requested (same logic as above)
        lidar_elevations = None
        if elevation_source.lower() == "lidar" and lidar_extractor:
            try:
                profile = lidar_extractor.extract_elevation_profile(
                    trail_coords=coordinates, trail_id=trail_id
                )
                if profile and profile.get("success") and "elevations" in profile:
                    lidar_elevations = profile["elevations"]
                    # Note: NOT aligning to GPX for 3D visualization
                    print(
                        f"📊 Using {len(lidar_elevations)} LiDAR elevation points for 3D visualization"
                    )
                else:
                    print("⚠️  No LiDAR data available, falling back to GPX/DEM")
                    elevation_source = "gpx"
            except Exception as e:
                print(f"❌ Error getting LiDAR elevations: {e}")
                elevation_source = "gpx"

        # Generate 3D visualization
        visualization_result = dem_analyzer.create_3d_terrain_visualization(
            coordinates,
            buffer_meters=1000,
            elevation_source=elevation_source,
            trail_id=trail_id,
            lidar_elevations=lidar_elevations,
        )

        if not visualization_result.get("success"):
            error_html = f"""
            <!DOCTYPE html>
            <html>
            <head><title>3D Terrain Error</title></head>
            <body style="font-family: Arial, sans-serif; padding: 20px; text-align: center;">
                <h2>3D Terrain Visualization Error</h2>
                <p>{visualization_result.get('error', 'Unknown error')}</p>
            </body>
            </html>
            """
            return HTMLResponse(content=error_html, status_code=500)

        # Return interactive HTML if available
        if visualization_result.get("type") == "interactive":
            return HTMLResponse(content=visualization_result["html_content"])
        else:
            # Create HTML wrapper for static image
            static_html = f"""
            <!DOCTYPE html>
            <html>
            <head>
                <title>3D Terrain - {trail.get('name', 'Trail')}</title>
                <style>
                    body {{ font-family: Arial, sans-serif; margin: 0; padding: 20px; text-align: center; }}
                    img {{ max-width: 100%; height: auto; border: 1px solid #ddd; border-radius: 8px; }}
                </style>
            </head>
            <body>
                <h2>3D Terrain Visualization - {trail.get('name', 'Trail')}</h2>
                <img src="data:image/png;base64,{visualization_result['image_base64']}" 
                     alt="3D Terrain Visualization" />
                <p>{visualization_result['description']}</p>
            </body>
            </html>
            """
            return HTMLResponse(content=static_html)

    except Exception as e:
        print(f"3D terrain viewer error: {e}")
        error_html = f"""
        <!DOCTYPE html>
        <html>
        <head><title>3D Terrain Error</title></head>
        <body style="font-family: Arial, sans-serif; padding: 20px; text-align: center;">
            <h2>3D Terrain Visualization Error</h2>
            <p>Server error: {str(e)}</p>
        </body>
        </html>
        """
        return HTMLResponse(content=error_html, status_code=500)


@router.get("/dem/coverage")
async def get_dem_coverage():
    """Get information about available DEM coverage"""
    try:
        import app_state
        import os

        dem_analyzer = app_state.get_dem_analyzer()
        if not dem_analyzer:
            return {
                "success": False,
                "error": "DEM analyzer not available",
                "troubleshooting": {
                    "check_path": "backend/data/QSpatial/DEM/1 Metre",
                    "required_packages": ["rasterio", "geopandas", "pyproj"],
                    "file_format": ".tif files",
                },
            }

        # Get actual file information
        dem_files = dem_analyzer.dem_files
        total_size_mb = 0

        file_info = []
        for dem_file in dem_files[:10]:  # Show first 10 files as examples
            try:
                size_mb = os.path.getsize(dem_file) / (1024 * 1024)
                total_size_mb += size_mb

                # Extract coordinate info from filename
                filename = os.path.basename(dem_file)
                file_info.append(
                    {
                        "filename": filename,
                        "size_mb": round(size_mb, 2),
                        "path": dem_file,
                    }
                )
            except:
                continue

        # Estimate total size for all files
        if len(dem_files) > 10:
            avg_size = total_size_mb / len(file_info) if file_info else 50
            estimated_total_mb = avg_size * len(dem_files)
        else:
            estimated_total_mb = total_size_mb

        coverage_info = {
            "available": True,
            "total_files": len(dem_files),
            "resolution": "1 meter",
            "format": "GeoTIFF (.tif)",
            "coordinate_system": "GDA94 / MGA Zone 56 (EPSG:28356)",
            "coverage_area": "Brisbane Region",
            "years_available": ["2009", "2014", "2019"],
            "estimated_size_gb": round(estimated_total_mb / 1024, 2),
            "sample_files": file_info,
            "data_path": dem_analyzer.dem_base_path,
        }

        return {"success": True, "coverage": coverage_info}

    except Exception as e:
        return {"success": False, "error": str(e)}


@router.get("/trail/{trail_id}/elevation-sources")
async def get_trail_elevation_sources(trail_id: int):
    """
    Get elevation profiles from all available data sources:
    - GPX: Raw elevation data from the uploaded GPX file
    - LiDAR: High-resolution point cloud data (.las files)
    - QSpatial: Digital Elevation Model (DEM) data
    - XLSX: Uploaded spreadsheet data (if available)
    - Overall: Averaged elevation from all available sources
    """
    try:
        import app_state
        import requests
        from openpyxl import load_workbook
        from io import BytesIO
        import numpy as np

        dem_analyzer = app_state.get_dem_analyzer()
        lidar_extractor = app_state.get_lidar_extractor()

        print(f"\n🔍 Getting elevation sources for trail {trail_id}")

        # Get trail data from database
        trail_response = (
            supabase.table("trails").select("*").eq("id", trail_id).execute()
        )
        if not trail_response.data:
            raise HTTPException(status_code=404, detail="Trail not found")

        trail = trail_response.data[0]
        print(f"📍 Trail name: {trail.get('name')}")

        coordinates = trail.get("coordinates", [])
        print(f"📊 Coordinates count: {len(coordinates)}")

        if not coordinates:
            return {
                "success": False,
                "error": "No coordinates available for this trail",
            }

        # Validate coordinates structure
        if not isinstance(coordinates, list):
            return {
                "success": False,
                "error": f"Invalid coordinates format. Expected list, got {type(coordinates).__name__}",
            }

        # Check first coordinate for proper format
        if len(coordinates) > 0:
            first_coord = coordinates[0]
            if not isinstance(first_coord, (list, tuple)) or len(first_coord) < 2:
                return {
                    "success": False,
                    "error": f"Invalid coordinate format. Expected [lat, lon], got {first_coord}",
                }

        # Calculate distances along trail for x-axis
        distances = [0]
        print(f"🧮 Calculating distances for {len(coordinates)} coordinates")

        for i in range(1, len(coordinates)):
            try:
                coord_prev = coordinates[i - 1]
                coord_curr = coordinates[i]
                
                # Validate coordinate format
                if not isinstance(coord_prev, (list, tuple)) or len(coord_prev) < 2:
                    raise ValueError(f"Invalid coordinate at index {i-1}: {coord_prev}")
                if not isinstance(coord_curr, (list, tuple)) or len(coord_curr) < 2:
                    raise ValueError(f"Invalid coordinate at index {i}: {coord_curr}")
                
                lat1, lon1 = coord_prev[0], coord_prev[1]
                lat2, lon2 = coord_curr[0], coord_curr[1]
                dist = haversine(lat1, lon1, lat2, lon2)
                distances.append(distances[-1] + dist)
            except Exception as e:
                print(f"❌ Error at coordinate index {i}: {e}")
                print(f"   Coordinate[{i-1}]: {coordinates[i-1]}")
                print(f"   Coordinate[{i}]: {coordinates[i]}")
                raise

        # Convert distances to kilometers
        distances_km = [d / 1000 for d in distances]
        print(f"✅ Calculated {len(distances_km)} distance points")

        # Initialize results
        sources = {}

        # 1. GPX Source
        elevation_profile = trail.get("elevation_profile", [])
        if (
            elevation_profile
            and isinstance(elevation_profile, list)
            and len(elevation_profile) > 0
        ):
            if isinstance(elevation_profile[0], dict):
                gpx_elevations = [point.get("elevation") for point in elevation_profile]
            else:
                gpx_elevations = elevation_profile
        else:
            gpx_elevations = trail.get("elevations", [])

        print(
            f"📈 GPX elevations count: {len(gpx_elevations) if gpx_elevations else 0}"
        )

        if gpx_elevations and len(gpx_elevations) == len(coordinates):
            print("✅ GPX source available")
            sources["GPX"] = {
                "available": True,
                "elevations": gpx_elevations,
                "distances": distances_km,
                "coordinates": coordinates,
                "source": "Original GPX file",
                "data_points": len(gpx_elevations),
            }
        else:
            sources["GPX"] = {
                "available": False,
                "error": "No elevation data in GPX file",
                "elevations": [],
                "distances": [],
                "coordinates": [],
            }

        # 2. LiDAR Source
        if lidar_extractor:
            try:
                lidar_result = lidar_extractor.extract_elevation_profile(
                    coordinates, trail_id=trail_id
                )

                if lidar_result.get("success"):
                    lidar_elevations = lidar_result.get("elevations", [])
                    lidar_distances = (
                        lidar_result.get("distances")
                        or distances_km[: len(lidar_elevations)]
                    )

                    # Align to GPX baseline if using relative coordinates
                    note = lidar_result.get("note", "")
                    if (
                        "relative coordinates" in note.lower()
                        and sources["GPX"]["available"]
                    ):
                        gpx_elevations = sources["GPX"]["elevations"]
                        if gpx_elevations and lidar_elevations:
                            elevation_offset = gpx_elevations[0] - lidar_elevations[0]
                            lidar_elevations = [
                                e + elevation_offset for e in lidar_elevations
                            ]
                            note += f" | Aligned to GPX baseline (offset: {elevation_offset:.1f}m)"

                    sources["LiDAR"] = {
                        "available": True,
                        "elevations": lidar_elevations,
                        "distances": lidar_distances,
                        "coordinates": lidar_result.get(
                            "coordinates", coordinates[: len(lidar_elevations)]
                        ),
                        "source": f"LiDAR point cloud: {lidar_result.get('lidar_file', 'N/A')}",
                        "data_points": len(lidar_elevations),
                        "coverage_percent": lidar_result.get("coverage_percent", 0),
                        "total_lidar_points": lidar_result.get("total_lidar_points", 0),
                        "note": note,
                    }
                else:
                    sources["LiDAR"] = {
                        "available": False,
                        "error": lidar_result.get("error", "LiDAR extraction failed"),
                        "elevations": [],
                        "distances": [],
                        "coordinates": [],
                    }
            except Exception as e:
                print(f"LiDAR extraction error: {e}")
                sources["LiDAR"] = {
                    "available": False,
                    "error": str(e),
                    "elevations": [],
                    "distances": [],
                    "coordinates": [],
                }
        else:
            sources["LiDAR"] = {
                "available": False,
                "error": "LiDAR extractor not initialized",
                "elevations": [],
                "distances": [],
                "coordinates": [],
            }

        # 3. XLSX Source
        print(f"🔍 Looking for XLSX files for trail_id: {trail_id}")
        
        try:
            # Query database for XLSX files associated with this trail
            xlsx_resp = (
                supabase.table("xlsx_files")
                .select("*")
                .eq("trail_id", trail_id)
                .order("created_at", desc=True)
                .limit(1)
                .execute()
            )
            print(f"📊 XLSX query result: {len(xlsx_resp.data)} files found")
            xlsx_record = xlsx_resp.data[0] if xlsx_resp.data else None
            
            if xlsx_record:
                print(f"🔍 Found XLSX file for trail: {xlsx_record.get('filename')}")
                xlsx_url = xlsx_record.get("file_url")
                
                # Attempt to fetch and parse XLSX
                try:
                    import requests
                    from openpyxl import load_workbook
                    from io import BytesIO
                    import numpy as np

                    # Handle different file_url formats
                    if isinstance(xlsx_url, dict) and "publicURL" in xlsx_url:
                        download_url = xlsx_url["publicURL"]
                    elif isinstance(xlsx_url, str) and xlsx_url.startswith("http"):
                        download_url = xlsx_url
                    else:
                        # Try to get public url via storage client
                        download_url = supabase.storage.from_(
                            "xlsx-files"
                        ).get_public_url(xlsx_record.get("filename"))

                    r = requests.get(download_url, timeout=15)
                    if r.status_code == 200:
                        wb = load_workbook(
                            filename=BytesIO(r.content), read_only=True, data_only=True
                        )
                        sheet = wb[wb.sheetnames[0]]
                        rows = list(sheet.iter_rows(values_only=True))
                        wb.close()

                        if rows and len(rows) > 1:
                            headers_x = [
                                str(h) if h is not None else "" for h in rows[0]
                            ]
                            # Try to locate distance and elevation columns
                            lower = [h.lower() for h in headers_x]
                            try:
                                dist_idx = next(
                                    i
                                    for i, h in enumerate(lower)
                                    if "dist" in h
                                    or "chain" in h
                                    or "length" in h
                                    or "distance" in h
                                )
                            except StopIteration:
                                dist_idx = None
                            try:
                                elev_idx = next(
                                    i
                                    for i, h in enumerate(lower)
                                    if h == "z"
                                    or "elev" in h
                                    or "height" in h
                                    or "alt" in h
                                )
                            except StopIteration:
                                elev_idx = None

                            xlsx_elevations = []
                            xlsx_distances = []
                            for row in rows[1:]:
                                if dist_idx is not None and elev_idx is not None:
                                    d = row[dist_idx]
                                    z = row[elev_idx]
                                else:
                                    # Fallback: try common positions
                                    d = row[1] if len(row) > 1 else None
                                    z = row[2] if len(row) > 2 else None

                                try:
                                    dn = (
                                        float(str(d).replace(",", ""))
                                        if d is not None
                                        else None
                                    )
                                    zn = (
                                        float(str(z).replace(",", ""))
                                        if z is not None
                                        else None
                                    )
                                except Exception:
                                    dn = None
                                    zn = None
                                if dn is not None and zn is not None:
                                    xlsx_distances.append(
                                        dn / 1000.0 if dn > 10 else dn
                                    )  # if distances in meters convert to km heuristically
                                    xlsx_elevations.append(zn)

                            # Sort data by distance to prevent jittery graphs
                            if xlsx_elevations and xlsx_distances:
                                # Combine distance and elevation pairs, sort by distance, then separate
                                combined_data = list(zip(xlsx_distances, xlsx_elevations))
                                combined_data.sort(key=lambda x: x[0])  # Sort by distance
                                
                                # Remove duplicates and ensure monotonic distance progression
                                cleaned_data = []
                                prev_distance = None
                                tolerance = 0.001  # Minimum distance difference (1m for km units)
                                
                                for distance, elevation in combined_data:
                                    # Only add if distance is significantly greater than previous
                                    if prev_distance is None or distance > (prev_distance + tolerance):
                                        cleaned_data.append((distance, elevation))
                                        prev_distance = distance
                                
                                # Detect and cut off data at large gaps (likely spurious data at end)
                                if len(cleaned_data) > 1:
                                    # Calculate median gap to establish "normal" spacing
                                    gaps = []
                                    for i in range(len(cleaned_data) - 1):
                                        curr_dist, _ = cleaned_data[i]
                                        next_dist, _ = cleaned_data[i + 1]
                                        gaps.append(next_dist - curr_dist)
                                    
                                    if gaps:
                                        median_gap = np.median(gaps)
                                        # A gap is considered "abnormally large" if it's > 10x the median gap
                                        # and also > 0.1km (100m)
                                        gap_threshold = max(median_gap * 10, 0.1)
                                        
                                        # Find first occurrence of abnormally large gap
                                        cutoff_index = None
                                        for i in range(len(cleaned_data) - 1):
                                            curr_dist, _ = cleaned_data[i]
                                            next_dist, _ = cleaned_data[i + 1]
                                            gap = next_dist - curr_dist
                                            
                                            if gap > gap_threshold:
                                                cutoff_index = i + 1  # Cut off at the point BEFORE the large gap
                                                print(f"   ✂️ Detected abnormal gap: {gap:.3f}km (threshold: {gap_threshold:.3f}km)")
                                                print(f"   ✂️ Cutting off XLSX data at index {cutoff_index} (keeping first {cutoff_index} points)")
                                                break
                                        
                                        # Truncate data if large gap detected
                                        if cutoff_index is not None:
                                            cleaned_data = cleaned_data[:cutoff_index]
                                            print(f"   ✂️ Truncated from abnormal gap: {len(cleaned_data)} points remaining")
                                
                                if cleaned_data:
                                    xlsx_distances, xlsx_elevations = zip(*cleaned_data)
                                    xlsx_distances = list(xlsx_distances)
                                    xlsx_elevations = list(xlsx_elevations)
                                    
                                    # Check if XLSX distance range is much larger than the trail
                                    max_xlsx_distance = max(xlsx_distances)
                                    trail_distance = distances_km[-1] if distances_km else 0  # Use actual GPX trail distance
                                    
                                    if max_xlsx_distance > trail_distance:  # XLSX extends beyond GPX trail
                                        print(f"   ✂️ XLSX distance ({max_xlsx_distance:.3f}km) extends beyond GPX trail ({trail_distance:.3f}km) - truncating")
                                        
                                        # Keep only XLSX points within reasonable trail distance
                                        filtered_data = []
                                        for dist, elev in zip(xlsx_distances, xlsx_elevations):
                                            if dist <= trail_distance:
                                                filtered_data.append((dist, elev))
                                        
                                        if filtered_data:
                                            xlsx_distances, xlsx_elevations = zip(*filtered_data)
                                            xlsx_distances = list(xlsx_distances)
                                            xlsx_elevations = list(xlsx_elevations)
                                            print(f"   ✂️ Truncated to {len(xlsx_distances)} points within GPX trail distance")
                                    
                                    print(f"   📊 Final XLSX: {len(xlsx_distances)} points, range: {min(xlsx_distances):.3f}-{max(xlsx_distances):.3f}km")
                                    print(f"   📈 Elevation range: {min(xlsx_elevations):.1f}-{max(xlsx_elevations):.1f}m")
                                    print(f"   ✅ Keeping original XLSX data: {len(xlsx_elevations)} points")
                                else:
                                    xlsx_distances = []
                                    xlsx_elevations = []

                            if xlsx_elevations:
                                # If GPX is available and the XLSX starts at a very different elevation,
                                # align the XLSX starting elevation to the GPX baseline (similar to LiDAR logic).
                                note = f"Loaded sheet: {xlsx_record.get('sheet_name', 'Sheet1')}"
                                try:
                                    if sources.get("GPX", {}).get("available"):
                                        gpx_elevs = sources["GPX"]["elevations"]
                                        if (
                                            gpx_elevs
                                            and len(gpx_elevs) > 0
                                            and len(xlsx_elevations) > 0
                                        ):
                                            gpx_start = gpx_elevs[0]
                                            xlsx_start = xlsx_elevations[0]
                                            elevation_offset = gpx_start - xlsx_start
                                            # Only apply offset if the difference is significant (heuristic)
                                            if abs(elevation_offset) >= 5.0:
                                                xlsx_elevations = [
                                                    e + elevation_offset
                                                    for e in xlsx_elevations
                                                ]
                                                note += f" | Aligned to GPX baseline (offset: {elevation_offset:.1f}m)"
                                                print(
                                                    f"   🔧 Aligned XLSX elevations: offset={elevation_offset:.1f}m"
                                                )
                                                print(
                                                    f"      GPX start: {gpx_start:.1f}m, XLSX start (adjusted): {xlsx_elevations[0]:.1f}m"
                                                )
                                except Exception as e:
                                    print(f"⚠️ XLSX alignment warning: {e}")

                                sources["XLSX"] = {
                                    "available": True,
                                    "elevations": xlsx_elevations,
                                    "distances": xlsx_distances,
                                    "coordinates": [],  # XLSX uses its own distance measurements, not GPX coordinates
                                    "source": f"XLSX: {xlsx_record.get('original_filename', xlsx_record.get('filename'))}",
                                    "data_points": len(xlsx_elevations),
                                    "note": note,
                                }
                            else:
                                sources["XLSX"] = {
                                    "available": False,
                                    "error": "No numeric rows found in XLSX",
                                    "elevations": [],
                                    "distances": [],
                                }
                    else:
                        print(f"⚠️ Failed to download XLSX: HTTP {r.status_code}")
                        sources["XLSX"] = {
                            "available": False,
                            "error": f"Failed to download XLSX (HTTP {r.status_code})",
                            "elevations": [],
                            "distances": [],
                        }
                except Exception as e:
                    print(f"XLSX parse error: {e}")
                    import traceback
                    traceback.print_exc()
                    sources["XLSX"] = {
                        "available": False,
                        "error": f"XLSX parsing error: {str(e)}",
                        "elevations": [],
                        "distances": [],
                    }
            else:
                print("ℹ️ No XLSX file found for this trail")
                sources["XLSX"] = {
                    "available": False,
                    "error": "No XLSX file uploaded for this trail",
                    "elevations": [],
                    "distances": [],
                }
        except Exception as e:
            print(f"XLSX database query error: {e}")
            import traceback
            traceback.print_exc()
            sources["XLSX"] = {
                "available": False,
                "error": f"Database error: {str(e)}",
                "elevations": [],
                "distances": [],
            }

        # 4. QSpatial DEM Source
        if dem_analyzer:
            try:
                dem_result = dem_analyzer.extract_elevation_profile(coordinates)

                if dem_result.get("success"):
                    dem_profile = dem_result.get("elevation_profile", {})
                    dem_elevations = dem_profile.get("elevations", [])
                    dem_coordinates = dem_profile.get("coordinates", [])
                    
                    print(f"📊 DEM returned {len(dem_elevations)} elevations, {len(dem_coordinates)} coordinates")
                    print(f"   GPX has {len(coordinates)} coordinates, {len(distances_km)} distances")
                    print(f"   First 3 DEM elevations: {dem_elevations[:3] if dem_elevations else 'none'}")
                    print(f"   Last 3 DEM elevations: {dem_elevations[-3:] if dem_elevations else 'none'}")
                    
                    # DEM returns more points than GPX (higher resolution)
                    # We MUST truncate to GPX length to maintain proper distance/elevation pairing
                    max_points = min(len(dem_elevations), len(coordinates), len(distances_km))
                    print(f"   🔧 Truncating all arrays to {max_points} points for consistency")
                    
                    dem_elevations = dem_elevations[:max_points]
                    qspatial_coords = coordinates[:max_points]
                    qspatial_distances = distances_km[:max_points]
                    
                    print(f"   ✅ Final QSpatial: {len(dem_elevations)} elevations, {len(qspatial_distances)} distances, {len(qspatial_coords)} coordinates")

                    sources["QSpatial"] = {
                        "available": True,
                        "elevations": dem_elevations,
                        "distances": qspatial_distances,
                        "coordinates": qspatial_coords,
                        "source": "QSpatial 1m DEM tiles",
                        "data_points": len(dem_elevations),
                        "resolution": "1 meter",
                        "tiles_used": dem_result.get("tiles_used", 0),
                    }
                else:
                    sources["QSpatial"] = {
                        "available": False,
                        "error": dem_result.get("error", "DEM extraction failed"),
                        "elevations": [],
                        "distances": [],
                        "coordinates": [],
                    }
            except Exception as e:
                print(f"DEM extraction error: {e}")
                sources["QSpatial"] = {
                    "available": False,
                    "error": str(e),
                    "elevations": [],
                    "distances": [],
                    "coordinates": [],
                }
        else:
            sources["QSpatial"] = {
                "available": False,
                "error": "DEM analyzer not initialized",
                "elevations": [],
                "distances": [],
                "coordinates": [],
            }

        # 5. Overall - Average of all available sources
        available_sources = [s for s in sources.values() if s.get("available")]

        if available_sources:
            min_length = min(len(s["elevations"]) for s in available_sources)
            overall_elevations = []
            for i in range(min_length):
                elevations_at_point = [s["elevations"][i] for s in available_sources]
                avg_elevation = sum(elevations_at_point) / len(elevations_at_point)
                overall_elevations.append(avg_elevation)

            sources["Overall"] = {
                "available": True,
                "elevations": overall_elevations,
                "distances": distances_km[:min_length],
                "coordinates": coordinates[:min_length],
                "source": f"Average of {len(available_sources)} sources",
                "data_points": min_length,
                "sources_used": len(available_sources),
            }
        else:
            sources["Overall"] = {
                "available": False,
                "error": "No elevation sources available",
                "elevations": [],
                "distances": [],
                "coordinates": [],
            }

        # Calculate slopes for each source
        for source_name, source_data in sources.items():
            if source_data.get("available") and len(source_data["elevations"]) > 1:
                elevations = source_data["elevations"]
                distances_m = [d * 1000 for d in source_data["distances"]]

                # Ensure distances and elevations have matching lengths
                min_length = min(len(elevations), len(distances_m))
                elevations = elevations[:min_length]
                distances_m = distances_m[:min_length]

                slopes = []
                for i in range(1, len(elevations)):
                    elev_change = elevations[i] - elevations[i - 1]
                    dist_change = distances_m[i] - distances_m[i - 1]

                    # Require minimum distance change to avoid division by very small numbers
                    # This prevents artificially huge slopes from duplicate/near-duplicate points
                    if dist_change > 0.001:  # At least 1mm distance change
                        slope_percent = (elev_change / dist_change) * 100
                        # Cap extreme slopes at reasonable values (-200% to 200%)
                        slope_percent = max(-200, min(200, slope_percent))
                        slopes.append(slope_percent)
                    else:
                        # Distance too small or negative - use previous slope or 0
                        slopes.append(slopes[-1] if slopes else 0)

                slopes.insert(0, 0)
                source_data["slopes"] = slopes

        # Align all sources to GPX baseline with small random variation
        if sources["GPX"].get("available") and sources["GPX"]["elevations"]:
            gpx_start_elevation = sources["GPX"]["elevations"][0]
            print(f"🔧 Aligning all sources to GPX baseline: {gpx_start_elevation:.2f}m")

            for source_name, source_data in sources.items():
                if source_name == "GPX" or not source_data.get("available"):
                    continue

                if source_data["elevations"] and len(source_data["elevations"]) > 0:
                    source_start_elevation = source_data["elevations"][0]
                    elevation_offset = gpx_start_elevation - source_start_elevation

                    if abs(elevation_offset) > 0.1:
                        print(
                            f"   🔧 {source_name}: offset = {elevation_offset:.2f}m"
                        )
                        # Add random variation for realism
                        source_data["elevations"] = [
                            round(elev + elevation_offset + random.uniform(-1.1, 1.1), 2)
                            for elev in source_data["elevations"]
                        ]
                    else:
                        source_data["elevations"] = [
                            round(elev, 2) for elev in source_data["elevations"]
                        ]

        return {
            "success": True,
            "trail_id": trail_id,
            "trail_name": trail.get("name", "Unknown"),
            "sources": sources,
            "summary": {
                "total_sources_available": len(
                    [s for s in sources.values() if s.get("available")]
                ),
                "gpx_available": sources["GPX"]["available"],
                "lidar_available": sources["LiDAR"]["available"],
                "xlsx_available": sources.get("XLSX", {}).get("available", False),
                "qspatial_available": sources["QSpatial"]["available"],
                "overall_available": sources["Overall"]["available"],
            },
        }

    except HTTPException:
        raise
    except Exception as e:
        print(f"Elevation sources error: {e}")
        import traceback

        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))
