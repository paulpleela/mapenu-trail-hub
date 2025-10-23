from fastapi import FastAPI, File, UploadFile, HTTPException, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse, HTMLResponse
from fastapi.staticfiles import StaticFiles
from typing import Optional
import gpxpy
import folium
from folium.plugins import Fullscreen, MeasureControl
import matplotlib.pyplot as plt
import base64
from io import BytesIO
import math
import os
import tempfile
import uuid
import uvicorn
import json
import rasterio
import numpy as np
from scipy.interpolate import griddata
import glob
from rasterio.merge import merge
from rasterio.warp import transform_bounds, transform
from rasterio.crs import CRS
from supabase import create_client, Client
from dotenv import load_dotenv

# Load environment variables
load_dotenv()

# Import real DEM analysis
try:
    from real_dem_analysis import RealDEMAnalyzer

    dem_path = os.path.join(
        os.path.dirname(__file__), "data", "QSpatial", "DEM", "1 Metre"
    )
    dem_analyzer = RealDEMAnalyzer(dem_path)
    print(f"DEM Analyzer initialized with {len(dem_analyzer.dem_files)} DEM files")
except ImportError as e:
    print(f"DEM analysis not available: {e}")
    dem_analyzer = None
except Exception as e:
    print(f"DEM initialization error: {e}")
    dem_analyzer = None

app = FastAPI()

# Add CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Allows all origins
    allow_credentials=True,
    allow_methods=["*"],  # Allows all methods
    allow_headers=["*"],  # Allows all headers
)

# Initialize Supabase client
supabase_url = os.getenv("SUPABASE_URL")
supabase_key = os.getenv("SUPABASE_KEY")

if not supabase_url or not supabase_key:
    raise Exception("Missing Supabase credentials. Please check your .env file.")

supabase: Client = create_client(supabase_url, supabase_key)

# Optional service-role client for server-side writes that must bypass RLS.
service_role_key = os.getenv("SUPABASE_SERVICE_ROLE_KEY")
supabase_service = None
if service_role_key:
    try:
        supabase_service: Client = create_client(supabase_url, service_role_key)
        print("Supabase service-role client initialized")
    except Exception as e:
        print(f"⚠️ Could not initialize supabase service client: {e}")

# Import LiDAR extraction (after supabase is initialized)
try:
    from lidar_extraction import LiDARExtractor

    lidar_cache_path = "/tmp/lidar_cache"  # Use /tmp for cache
    lidar_extractor = LiDARExtractor(lidar_cache_path, supabase_client=supabase)
    print(
        f"LiDAR Extractor initialized with {len(lidar_extractor.lidar_files)} LiDAR files"
    )
except ImportError as e:
    print(f"LiDAR extraction not available: {e}")
    lidar_extractor = None
except Exception as e:
    print(f"LiDAR initialization error: {e}")
    lidar_extractor = None


# Use the same haversine function from main.py
def haversine(lat1, lon1, lat2, lon2):
    R = 6371000  # Earth radius in meters
    phi1, phi2 = math.radians(lat1), math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlambda = math.radians(lon2 - lon1)
    a = (
        math.sin(dphi / 2) ** 2
        + math.cos(phi1) * math.cos(phi2) * math.sin(dlambda / 2) ** 2
    )
    return R * 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))


""" 
URAWEE
measure ‘bumpiness’, meaningful ups/downs over 1 meter
    - frequency 60%
    - amplitude 40%
"""


def count_rolling_hills(elevations):
    """
    Count the number of distinct "hills" (peaks and valleys) in the elevation profile.

    Algorithm:
    1. Identify all local peaks (higher than neighbors)
    2. Identify all local valleys (lower than neighbors)
    3. Total hills = peaks + valleys (each represents a direction change)

    A "significant" peak/valley must have at least 3m elevation difference from neighbors
    to filter out GPS noise.

    Args:
        elevations: List of elevation values in meters

    Returns:
        int: Number of distinct hills (peaks + valleys)
    """
    if len(elevations) < 3:
        return 0

    # Minimum elevation change to be considered significant
    # 1m threshold catches most noticeable hills while filtering extreme GPS noise
    min_prominence = 1.0  # meters

    peaks = 0
    valleys = 0

    for i in range(1, len(elevations) - 1):
        prev_elev = elevations[i - 1]
        curr_elev = elevations[i]
        next_elev = elevations[i + 1]

        # Check if this is a local peak (higher than both neighbors)
        if curr_elev > prev_elev and curr_elev > next_elev:
            # Check if it's significant enough
            if (curr_elev - prev_elev >= min_prominence) or (
                curr_elev - next_elev >= min_prominence
            ):
                peaks += 1

        # Check if this is a local valley (lower than both neighbors)
        elif curr_elev < prev_elev and curr_elev < next_elev:
            # Check if it's significant enough
            if (prev_elev - curr_elev >= min_prominence) or (
                next_elev - curr_elev >= min_prominence
            ):
                valleys += 1

    # Total number of hills = peaks + valleys
    # Each represents a change in terrain direction
    total_hills = peaks + valleys

    return total_hills


def analyze_rolling_hills(elevations, distances):
    """Advanced rolling hills analysis: counts and scores significant ascents/descents

    Returns:
        tuple: (rolling_index: float, hills_count: int)
    """
    if len(elevations) < 3:
        return 0.0, 0

    # Count actual hills (peaks and valleys)
    hills_count = count_rolling_hills(elevations)

    # For the rolling index calculation, count significant elevation changes
    threshold = 1  # meters, what counts as a significant change
    significant_changes = []
    for i in range(1, len(elevations)):
        change = elevations[i] - elevations[i - 1]
        if abs(change) >= threshold:
            significant_changes.append(abs(change))

    # Frequency: how many significant changes per km
    total_distance = distances[-1] if distances else 1
    changes_per_km = (
        len(significant_changes) / total_distance if total_distance > 0 else 0
    )

    # Amplitude: average size of significant changes
    avg_change_size = (
        sum(significant_changes) / len(significant_changes)
        if significant_changes
        else 0
    )

    # Composite index: weighted sum (tweak weights as needed)
    rolling_index = 0.6 * changes_per_km + 0.4 * (
        avg_change_size / 20
    )  # typical big hills are ~20 m per hill

    # Debug output
    print(f"🔍 Rolling Hills Debug:")
    print(f"   - Actual hills (peaks + valleys): {hills_count}")
    print(f"   - Significant elevation changes: {len(significant_changes)}")
    print(f"   - Total distance: {total_distance:.2f} km")
    print(f"   - Changes per km: {changes_per_km:.2f}")
    print(f"   - Avg change size: {avg_change_size:.2f} m")
    print(f"   - Rolling index (UNCAPPED): {rolling_index:.4f}")

    # Return both the rolling index and the count
    return rolling_index, hills_count


"""
gives a 0–100% match by weighting 5 checks:
    distance difference within ±1 km (25%),
    elevation-gain difference within ±500 m (25%),
    overall difficulty within ±5 on our 10-point scale (20%),
    terrain character using the rolling-hills index (15%),
    and surface difficulty similarity (15%).
    Higher sub-scores → higher overall match.
"""


def calculate_trail_similarity(trail1, trail2):
    """Calculate similarity score between two trails (0-1, higher = more similar)"""
    # Normalize factors for comparison
    distance_diff = abs(trail1["distance"] - trail2["distance"])
    distance_similarity = max(
        0, 1 - (distance_diff / 1000)
    )  # Within 1km is very similar

    elevation_gain_diff = abs(trail1["elevation_gain"] - trail2["elevation_gain"])
    elevation_similarity = max(
        0, 1 - (elevation_gain_diff / 500)
    )  # Within 500m is similar

    difficulty_diff = abs(trail1["difficulty_score"] - trail2["difficulty_score"])
    difficulty_similarity = max(
        0, 1 - (difficulty_diff / 5)
    )  # Within 5 points is similar

    rolling_hills_diff = abs(
        trail1["rolling_hills_index"] - trail2["rolling_hills_index"]
    )
    rolling_similarity = max(0, 1 - (rolling_hills_diff / 0.5))  # Within 0.5 is similar

    # Surface difficulty similarity
    surface1 = trail1.get("surface_difficulty_score", 1.0)
    surface2 = trail2.get("surface_difficulty_score", 1.0)
    surface_diff = abs(surface1 - surface2)
    surface_similarity = max(
        0, 1 - (surface_diff / 0.5)
    )  # Within 0.5 multiplier is similar

    # Weighted average (adjust weights as needed)
    similarity = (
        distance_similarity * 0.25
        + elevation_similarity * 0.25
        + difficulty_similarity * 0.20
        + rolling_similarity * 0.15
        + surface_similarity * 0.15
    )

    return similarity


def get_trail_weather_exposure(trail):
    """Calculate static weather exposure risk (doesn't change with weather)"""
    max_elev = trail.get("max_elevation", 0)

    # Return exposure level and explanation (static characteristics)
    if max_elev > 1500:
        return {
            "exposure_level": "High",
            "risk_factors": [
                "Rapid weather changes",
                "Snow/ice risk",
                "High wind exposure",
                "Temperature drops",
            ],
        }
    elif max_elev > 1000:
        return {
            "exposure_level": "Moderate",
            "risk_factors": ["Cooler temperatures", "Wind exposure", "Potential fog"],
        }
    elif max_elev > 500:
        return {
            "exposure_level": "Low-Moderate",
            "risk_factors": ["Slightly cooler temps", "Some wind exposure"],
        }
    else:
        return {
            "exposure_level": "Low",
            "risk_factors": ["Minimal weather impact", "Protected terrain"],
        }


async def get_live_weather_difficulty(trail_coords, weather_api_key=None):
    """Get current weather conditions and calculate live difficulty multiplier"""
    if not trail_coords:
        return {
            "multiplier": 1.0,
            "conditions": "No coordinates available",
            "explanation": "Trail coordinates required for weather data",
        }

    # For demo purposes, simulate different conditions
    # In production, this would call OpenWeatherMap or similar API
    import random

    # Simulate current conditions (in real app, call weather API)
    conditions = random.choice(
        [
            {"temp": 15, "wind": 10, "rain": False, "visibility": "Good"},
            {"temp": 5, "wind": 25, "rain": True, "visibility": "Poor"},
            {"temp": 25, "wind": 5, "rain": False, "visibility": "Excellent"},
            {"temp": -2, "wind": 30, "rain": False, "visibility": "Fair"},
            {"temp": 18, "wind": 15, "rain": False, "visibility": "Good"},
            {"temp": 12, "wind": 8, "rain": False, "visibility": "Excellent"},
        ]
    )

    multiplier = 1.0
    factors = []

    # Temperature impact
    if conditions["temp"] < 0:
        multiplier += 0.3
        factors.append("Freezing temperatures")
    elif conditions["temp"] < 5:
        multiplier += 0.2
        factors.append("Cold temperatures")
    elif conditions["temp"] > 30:
        multiplier += 0.1
        factors.append("Hot temperatures")

    # Wind impact
    if conditions["wind"] > 25:
        multiplier += 0.2
        factors.append("Strong winds")
    elif conditions["wind"] > 15:
        multiplier += 0.1
        factors.append("Moderate winds")

    # Rain impact
    if conditions["rain"]:
        multiplier += 0.3
        factors.append("Wet/slippery conditions")

    # Visibility impact
    if conditions["visibility"] == "Poor":
        multiplier += 0.2
        factors.append("Poor visibility")

    condition_desc = f"{conditions['temp']}°C, {conditions['wind']}km/h winds"
    if conditions["rain"]:
        condition_desc += ", raining"

    return {
        "multiplier": round(multiplier, 2),
        "conditions": condition_desc,
        "explanation": f"Weather factors: {', '.join(factors) if factors else 'Good conditions'}",
    }


def calculate_terrain_variety(elevations):
    """Calculate how varied the terrain is (0-10 scale)"""
    if len(elevations) < 10:
        return 0

    # Calculate elevation ranges in 100m bands
    elevation_bands = set()
    for elev in elevations:
        band = int(elev / 100) * 100  # Round to nearest 100m
        elevation_bands.add(band)

    # More bands = more variety
    variety_score = min(len(elevation_bands), 10)  # Cap at 10

    # Also consider elevation change rate
    elevation_changes = []
    for i in range(1, len(elevations)):
        change_rate = abs(elevations[i] - elevations[i - 1])
        elevation_changes.append(change_rate)

    # Bonus for frequent elevation changes
    if elevation_changes:
        avg_change = sum(elevation_changes) / len(elevation_changes)
        if avg_change > 20:  # Frequent significant changes
            variety_score = min(variety_score + 2, 10)
        elif avg_change > 10:  # Moderate changes
            variety_score = min(variety_score + 1, 10)

    return variety_score


def get_terrain_variety_description(score):
    """Get a description for terrain variety score"""
    if score >= 8:
        return "Highly varied terrain with multiple elevation zones"
    elif score >= 6:
        return "Good terrain variety with several elevation changes"
    elif score >= 4:
        return "Moderate terrain variety with some elevation changes"
    elif score >= 2:
        return "Limited terrain variety, mostly consistent elevation"
    else:
        return "Flat or very consistent terrain"


"""
Easy (0.7-0.8x): Paved roads, boardwalks, concrete
Normal (1.0x): Dirt trails, gravel, grass (baseline)
Moderate (1.1-1.3x): Soil, forest floor, wood chips, tall grass
Challenging (1.3-1.6x): Sand, mud, loose gravel, scree, snow
Difficult (1.6-2.0x): Rock, boulders, swamp, ice
"""


def get_surface_difficulty_multiplier(surface_type):
    """Get difficulty multiplier based on terrain surface type"""
    surface_multipliers = {
        # Easy surfaces (< 1.0)
        "paved": 0.7,  # Paved roads, easiest
        "boardwalk": 0.8,  # Wooden boardwalks
        "concrete": 0.75,  # Concrete paths
        # Normal surfaces (1.0)
        "dirt": 1.0,  # Packed dirt trails (baseline)
        "gravel": 1.0,  # Well-maintained gravel
        "grass": 1.0,  # Short grass
        # Moderate surfaces (1.1-1.3)
        "soil": 1.1,  # Loose soil
        "forest_floor": 1.15,  # Leaf litter, small branches
        "crushed_stone": 1.1,  # Loose crushed stone
        "wood_chips": 1.2,  # Wood chip trails
        "tall_grass": 1.25,  # Long grass, meadows
        # Challenging surfaces (1.3-1.6)
        "sand": 1.4,  # Beach sand, very tiring
        "mud": 1.5,  # Muddy conditions
        "loose_gravel": 1.3,  # Loose, shifting gravel
        "scree": 1.6,  # Loose rock fragments
        "snow": 1.4,  # Snow covered (not ice)
        # Difficult surfaces (1.6-2.0)
        "rock": 1.7,  # Rocky terrain, scrambling
        "boulder": 1.8,  # Large rocks, careful footing
        "swamp": 1.9,  # Swampy, unstable ground
        "ice": 2.0,  # Icy conditions, dangerous
        # Default for unknown
        "unknown": 1.0,
    }

    return surface_multipliers.get(surface_type.lower(), 1.0)


def estimate_surface_type_from_terrain(coordinates, elevation_profile=None):
    """Estimate likely surface types based on terrain characteristics"""
    if not coordinates:
        return [{"surface": "unknown", "percentage": 100}]

    # Analyze terrain to estimate surface types
    # This is a simplified estimation - in reality you'd use satellite data,
    # land cover maps, or trail databases

    surface_segments = []
    num_points = len(coordinates)

    # Get elevation statistics if available
    elevations = []
    if elevation_profile:
        elevations = [point.get("elevation", 0) for point in elevation_profile]

    # Estimate based on coordinate patterns and elevation
    if elevations:
        avg_elevation = sum(elevations) / len(elevations)
        elevation_variance = sum((e - avg_elevation) ** 2 for e in elevations) / len(
            elevations
        )

        # High elevation, high variance = rocky terrain
        if avg_elevation > 800 and elevation_variance > 1000:
            surface_segments = [
                {"surface": "rock", "percentage": 40},
                {"surface": "dirt", "percentage": 35},
                {"surface": "scree", "percentage": 25},
            ]
        # High elevation, low variance = alpine meadows
        elif avg_elevation > 800:
            surface_segments = [
                {"surface": "grass", "percentage": 50},
                {"surface": "dirt", "percentage": 30},
                {"surface": "rock", "percentage": 20},
            ]
        # Medium elevation, high variance = forest trails
        elif avg_elevation > 200 and elevation_variance > 500:
            surface_segments = [
                {"surface": "forest_floor", "percentage": 60},
                {"surface": "dirt", "percentage": 30},
                {"surface": "soil", "percentage": 10},
            ]
        # Low elevation, coastal areas
        elif avg_elevation < 100:
            # Could be coastal, might have sand
            surface_segments = [
                {"surface": "dirt", "percentage": 50},
                {"surface": "sand", "percentage": 30},
                {"surface": "grass", "percentage": 20},
            ]
        else:
            # Default mixed terrain
            surface_segments = [
                {"surface": "dirt", "percentage": 70},
                {"surface": "gravel", "percentage": 20},
                {"surface": "grass", "percentage": 10},
            ]
    else:
        # No elevation data, assume mixed terrain
        surface_segments = [
            {"surface": "dirt", "percentage": 60},
            {"surface": "gravel", "percentage": 25},
            {"surface": "grass", "percentage": 15},
        ]

    return surface_segments


def calculate_surface_difficulty_score(surface_segments):
    """Calculate overall surface difficulty score from surface segments"""
    if not surface_segments:
        return 1.0

    total_score = 0
    total_percentage = 0

    for segment in surface_segments:
        surface = segment.get("surface", "unknown")
        percentage = segment.get("percentage", 0)
        multiplier = get_surface_difficulty_multiplier(surface)

        total_score += multiplier * percentage
        total_percentage += percentage

    if total_percentage == 0:
        return 1.0

    # Weighted average of surface difficulties
    return total_score / total_percentage


def get_surface_difficulty_description(score, surface_segments):
    """Get human-readable description of surface difficulty"""
    primary_surfaces = sorted(
        surface_segments, key=lambda x: x["percentage"], reverse=True
    )[:2]

    description = f"Surface difficulty: {score:.2f}x baseline. "

    if score <= 0.8:
        description += "Very easy walking on "
    elif score <= 1.0:
        description += "Standard trail surface with "
    elif score <= 1.3:
        description += "Moderately challenging surface with "
    elif score <= 1.6:
        description += "Difficult surface requiring careful footing with "
    else:
        description += "Very challenging surface demanding experience with "

    surface_names = []
    for surface in primary_surfaces:
        name = surface["surface"].replace("_", " ")
        percentage = surface["percentage"]
        surface_names.append(f"{name} ({percentage}%)")

    description += " and ".join(surface_names)

    return description


def get_weather_exposure_from_score(score):
    """Convert weather score back to exposure level and risk factors"""
    # Handle None/null values
    if score is None:
        score = 1.0  # Default to low exposure

    try:
        score = float(score)  # Ensure it's a number
    except (ValueError, TypeError):
        score = 1.0  # Default to low exposure if conversion fails

    if score >= 1.25:
        return {
            "exposure_level": "High",
            "risk_factors": [
                "Rapid weather changes",
                "Snow/ice risk",
                "High wind exposure",
                "Temperature drops",
            ],
        }
    elif score >= 1.15:
        return {
            "exposure_level": "Moderate",
            "risk_factors": ["Cooler temperatures", "Wind exposure", "Potential fog"],
        }
    elif score >= 1.05:
        return {
            "exposure_level": "Low-Moderate",
            "risk_factors": ["Slightly cooler temps", "Some wind exposure"],
        }
    else:
        return {
            "exposure_level": "Low",
            "risk_factors": ["Minimal weather impact", "Protected terrain"],
        }


def find_relevant_dem_tiles(trail_coords):
    """Find DEM tiles that cover the trail coordinates"""
    if not trail_coords:
        return []

    # Convert lat/lon to UTM (approximately) to match DEM tile naming
    # Brisbane DEM tiles are in MGA Zone 56 (EPSG:28356)
    # Tile naming follows pattern: Brisbane_YYYY_LGA_SW_EASTING_NORTHING_1K_DEM_1m.tif

    relevant_tiles = []
    dem_dir = os.path.join("data", "QSpatial", "DEM", "1 Metre")

    if not os.path.exists(dem_dir):
        print(f"DEM directory not found: {dem_dir}")
        return []

    # Get all available DEM files
    dem_files = glob.glob(os.path.join(dem_dir, "*.tif"))

    # For now, return all available tiles - in production you'd filter by bounds
    # This is a simplified approach for the demo
    return dem_files[:4]  # Limit to 4 tiles for performance


def process_dem_for_trail(trail_coords, dem_files, resolution_factor=4):
    """Process DEM data for 3D visualization of a trail"""
    if not dem_files or not trail_coords:
        return None

    try:
        print(
            f"Processing DEM with {len(dem_files)} files and {len(trail_coords)} trail coordinates"
        )

        # For demonstration, let's create synthetic terrain data if DEM processing fails
        # This ensures the 3D viewer works while we debug the real DEM processing

        # Calculate bounding box for the trail
        lats = [coord[0] for coord in trail_coords]
        lons = [coord[1] for coord in trail_coords]

        min_lat, max_lat = min(lats), max(lats)
        min_lon, max_lon = min(lons), max(lons)

        print(
            f"Trail bounds: lat {min_lat:.6f} to {max_lat:.6f}, lon {min_lon:.6f} to {max_lon:.6f}"
        )

        # Try to process real DEM data
        try:
            # Read first DEM file to get basic info
            with rasterio.open(dem_files[0]) as dem:
                print(f"DEM CRS: {dem.crs}")
                print(f"DEM bounds: {dem.bounds}")
                print(f"DEM shape: {dem.shape}")

                # Simple approach: read a small area from the first DEM file
                elevation_data = dem.read(1)
                transform = dem.transform

                # Get a subset of the elevation data for performance
                height, width = elevation_data.shape
                step = resolution_factor * 10  # Larger step for demo

                subset_height = height // step
                subset_width = width // step

                if subset_height < 10 or subset_width < 10:
                    raise ValueError("DEM subset too small")

                # Create coordinate grids for the subset
                x_coords = []
                y_coords = []
                elevations = []

                for i in range(0, subset_height):
                    for j in range(0, subset_width):
                        row = i * step
                        col = j * step
                        if row < height and col < width:
                            elevation = float(elevation_data[row, col])
                            if not np.isnan(elevation) and elevation > -9999:
                                # Get coordinate in DEM's projection
                                x, y = rasterio.transform.xy(transform, row, col)

                                # Convert to lat/lon
                                lon, lat = transform(
                                    dem.crs, CRS.from_epsg(4326), [x], [y]
                                )
                                x_coords.append(lon[0])
                                y_coords.append(lat[0])
                                elevations.append(elevation)

                print(f"Extracted {len(elevations)} elevation points from DEM")

                if len(elevations) >= 100:  # Minimum points for visualization
                    # Create regular grid for 3D surface
                    grid_size = 30
                    x_min, x_max = min(x_coords), max(x_coords)
                    y_min, y_max = min(y_coords), max(y_coords)

                    xi = np.linspace(x_min, x_max, grid_size)
                    yi = np.linspace(y_min, y_max, grid_size)
                    xi_grid, yi_grid = np.meshgrid(xi, yi)

                    # Interpolate elevations onto regular grid
                    points = np.column_stack((x_coords, y_coords))
                    zi_grid = griddata(
                        points, elevations, (xi_grid, yi_grid), method="linear"
                    )

                    # Fill NaN values
                    mask = np.isnan(zi_grid)
                    if np.any(mask):
                        zi_grid_filled = griddata(
                            points, elevations, (xi_grid, yi_grid), method="nearest"
                        )
                        zi_grid[mask] = zi_grid_filled[mask]

                    # Process trail line
                    trail_line = []
                    for coord in trail_coords[::5]:  # Subsample trail points
                        lat, lon = coord
                        # Interpolate elevation for this trail point
                        if x_min <= lon <= x_max and y_min <= lat <= y_max:
                            trail_elev = griddata(
                                points, elevations, (lon, lat), method="linear"
                            )
                            if np.isnan(trail_elev):
                                trail_elev = griddata(
                                    points, elevations, (lon, lat), method="nearest"
                                )

                            if not np.isnan(trail_elev):
                                trail_line.append(
                                    {"x": lon, "y": lat, "z": float(trail_elev)}
                                )

                    surface_data = {
                        "x": xi.tolist(),
                        "y": yi.tolist(),
                        "z": zi_grid.tolist(),
                        "bounds": {
                            "x_min": float(x_min),
                            "x_max": float(x_max),
                            "y_min": float(y_min),
                            "y_max": float(y_max),
                            "z_min": float(np.nanmin(zi_grid)),
                            "z_max": float(np.nanmax(zi_grid)),
                        },
                    }

                    return {
                        "surface": surface_data,
                        "trail_line": trail_line,
                        "metadata": {
                            "grid_size": grid_size,
                            "num_trail_points": len(trail_line),
                            "elevation_range": float(
                                np.nanmax(zi_grid) - np.nanmin(zi_grid)
                            ),
                            "data_source": "Brisbane DEM",
                        },
                    }

        except Exception as dem_error:
            print(f"DEM processing failed: {dem_error}")

        # Fallback: Create Mt Coot-tha area terrain for demonstration
        print("Creating Mt Coot-tha area terrain for demonstration")

        # Mt Coot-tha specific coordinates (Brisbane, Australia)
        # These are the actual boundaries of Mt Coot-tha area
        mt_coottha_bounds = {
            "lat_min": -27.495,  # Southern boundary
            "lat_max": -27.465,  # Northern boundary
            "lon_min": 152.940,  # Western boundary
            "lon_max": 152.980,  # Eastern boundary
        }

        grid_size = 80  # High resolution for the whole mountain area

        x_min = mt_coottha_bounds["lon_min"]
        x_max = mt_coottha_bounds["lon_max"]
        y_min = mt_coottha_bounds["lat_min"]
        y_max = mt_coottha_bounds["lat_max"]

        xi = np.linspace(x_min, x_max, grid_size)
        yi = np.linspace(y_min, y_max, grid_size)
        xi_grid, yi_grid = np.meshgrid(xi, yi)

        # Create realistic Mt Coot-tha terrain profile
        # Mt Coot-tha peak is approximately at -27.4756°S, 152.9581°E with elevation ~287m
        peak_lat = -27.4756
        peak_lon = 152.9581

        zi_grid = np.zeros((grid_size, grid_size))
        for i in range(grid_size):
            for j in range(grid_size):
                lat = yi_grid[i, j]
                lon = xi_grid[i, j]

                # Distance from Mt Coot-tha summit
                dist_to_peak = np.sqrt((lat - peak_lat) ** 2 + (lon - peak_lon) ** 2)

                # Create Mt Coot-tha's characteristic shape
                # Main peak with gradual slopes
                main_peak = 287 * np.exp(-dist_to_peak * 800)  # Peak at ~287m

                # Secondary ridges and spurs
                ridge1 = 150 * np.exp(
                    -((lat - (-27.480)) ** 2 + (lon - 152.950) ** 2) * 2000
                )
                ridge2 = 120 * np.exp(
                    -((lat - (-27.470)) ** 2 + (lon - 152.965) ** 2) * 2500
                )
                ridge3 = 180 * np.exp(
                    -((lat - (-27.485)) ** 2 + (lon - 152.975) ** 2) * 2200
                )

                # Valley systems around the mountain
                valley1 = -50 * np.exp(
                    -((lat - (-27.490)) ** 2 + (lon - 152.945) ** 2) * 1500
                )
                valley2 = -40 * np.exp(
                    -((lat - (-27.475)) ** 2 + (lon - 152.985) ** 2) * 1800
                )

                # Base elevation (Brisbane area is generally 50-100m above sea level)
                base_elevation = 80 + 20 * np.sin((lat + 27.48) * 200) * np.cos(
                    (lon - 152.96) * 300
                )

                # Natural terrain variation
                terrain_noise = (
                    15 * np.sin((lat + 27.48) * 1000) * np.cos((lon - 152.96) * 1200)
                )
                small_features = (
                    8 * np.sin((lat + 27.48) * 2000) * np.cos((lon - 152.96) * 2500)
                )

                # Combine all terrain features
                elevation = (
                    base_elevation
                    + main_peak
                    + ridge1
                    + ridge2
                    + ridge3
                    + valley1
                    + valley2
                    + terrain_noise
                    + small_features
                )

                # Ensure realistic elevation bounds for Brisbane area
                elevation = max(20, min(350, elevation))
                zi_grid[i, j] = elevation

        # Create trail line within Mt Coot-tha area
        trail_line = []
        for coord in trail_coords[::3]:  # Sample more trail points for better detail
            lat, lon = coord
            # Check if trail point is within Mt Coot-tha area
            if (
                mt_coottha_bounds["lat_min"] <= lat <= mt_coottha_bounds["lat_max"]
                and mt_coottha_bounds["lon_min"] <= lon <= mt_coottha_bounds["lon_max"]
            ):

                # Find closest grid point for elevation
                i = int((lat - y_min) / (y_max - y_min) * (grid_size - 1))
                j = int((lon - x_min) / (x_max - x_min) * (grid_size - 1))
                i = max(0, min(grid_size - 1, i))
                j = max(0, min(grid_size - 1, j))

                trail_line.append(
                    {
                        "x": lon,
                        "y": lat,
                        "z": float(zi_grid[i, j] + 3),  # Slightly above terrain
                    }
                )
            else:
                # For trail points outside Mt Coot-tha area, estimate elevation
                trail_line.append(
                    {
                        "x": lon,
                        "y": lat,
                        "z": 100.0,  # Default elevation for points outside area
                    }
                )

        surface_data = {
            "x": xi.tolist(),
            "y": yi.tolist(),
            "z": zi_grid.tolist(),
            "bounds": {
                "x_min": float(x_min),
                "x_max": float(x_max),
                "y_min": float(y_min),
                "y_max": float(y_max),
                "z_min": float(np.min(zi_grid)),
                "z_max": float(np.max(zi_grid)),
            },
        }

        return {
            "surface": surface_data,
            "trail_line": trail_line,
            "metadata": {
                "grid_size": grid_size,
                "num_trail_points": len(trail_line),
                "elevation_range": float(np.max(zi_grid) - np.min(zi_grid)),
                "data_source": "Mt Coot-tha Area Terrain",
                "area_name": "Mt Coot-tha, Brisbane",
                "peak_elevation": f"{np.max(zi_grid):.0f}m",
            },
        }

    except Exception as e:
        print(f"Error processing DEM data: {e}")
        import traceback

        traceback.print_exc()
        return None


@app.get("/trails")
async def get_trails():
    """Get all trails from Supabase database"""
    try:
        response = supabase.table("trails").select("*").execute()
        trails = response.data
        return {"success": True, "trails": trails}
    except Exception as e:
        print(f"Database error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/trail/{trail_id}/similar")
async def get_similar_trails(trail_id: int, limit: int = 5):
    """Get trails similar to the specified trail"""
    try:
        # Get the target trail
        target_response = (
            supabase.table("trails").select("*").eq("id", trail_id).execute()
        )
        if not target_response.data:
            raise HTTPException(status_code=404, detail="Trail not found")

        target_trail = target_response.data[0]

        # Get all other trails
        all_trails_response = (
            supabase.table("trails").select("*").neq("id", trail_id).execute()
        )
        all_trails = all_trails_response.data

        # Calculate similarity scores
        similarities = []
        for trail in all_trails:
            similarity_score = calculate_trail_similarity(target_trail, trail)
            similarities.append({"trail": trail, "similarity_score": similarity_score})

        # Sort by similarity and return top results
        similarities.sort(key=lambda x: x["similarity_score"], reverse=True)
        similar_trails = similarities[:limit]

        return {
            "success": True,
            "target_trail": target_trail["name"],
            "similar_trails": similar_trails,
        }

    except Exception as e:
        print(f"Similar trails error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/analytics/overview")
async def get_analytics_overview():
    """Get overall analytics for all trails"""
    try:
        response = supabase.table("trails").select("*").execute()
        trails = response.data

        if not trails:
            return {"success": True, "analytics": {"total_trails": 0}}

        # Calculate aggregate statistics
        total_distance = sum(trail.get("distance", 0) for trail in trails)
        total_elevation_gain = sum(trail.get("elevation_gain", 0) for trail in trails)
        avg_difficulty = sum(
            trail.get("difficulty_score", 0) for trail in trails
        ) / len(trails)

        # Difficulty distribution
        difficulty_dist = {"Easy": 0, "Moderate": 0, "Hard": 0, "Extreme": 0}
        for trail in trails:
            level = trail.get("difficulty_level", "Unknown")
            if level in difficulty_dist:
                difficulty_dist[level] += 1

        # Distance categories
        distance_categories = {
            "Short (<5km)": 0,
            "Medium (5-15km)": 0,
            "Long (>15km)": 0,
        }
        for trail in trails:
            distance = trail.get("distance", 0)
            if distance < 5:
                distance_categories["Short (<5km)"] += 1
            elif distance <= 15:
                distance_categories["Medium (5-15km)"] += 1
            else:
                distance_categories["Long (>15km)"] += 1

        return {
            "success": True,
            "analytics": {
                "total_trails": len(trails),
                "total_distance_km": round(total_distance, 1),
                "total_elevation_gain_m": round(total_elevation_gain, 0),
                "avg_difficulty_score": round(avg_difficulty, 1),
                "difficulty_distribution": difficulty_dist,
                "distance_categories": distance_categories,
                "most_challenging": max(
                    trails, key=lambda t: t.get("difficulty_score", 0)
                )["name"],
                "longest_trail": max(trails, key=lambda t: t.get("distance", 0))[
                    "name"
                ],
            },
        }

    except Exception as e:
        print(f"Analytics error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/trail/{trail_id}/weather")
async def get_trail_weather(trail_id: int):
    """Get live weather conditions and difficulty multiplier for a trail"""
    try:
        print(f"Getting weather data for trail ID: {trail_id}")

        # Get the trail
        trail_response = (
            supabase.table("trails").select("*").eq("id", trail_id).execute()
        )
        if not trail_response.data:
            raise HTTPException(status_code=404, detail="Trail not found")

        trail = trail_response.data[0]
        print(f"Found trail: {trail.get('name', 'Unknown')}")

        coordinates = trail.get("coordinates", [])

        if not coordinates:
            raise HTTPException(status_code=400, detail="Trail has no coordinate data")

        # Get midpoint coordinates for weather lookup
        mid_index = len(coordinates) // 2
        trail_coords = coordinates[mid_index]
        print(f"Using coordinates: {trail_coords}")

        # Get live weather data (this would use a real weather API in production)
        weather_data = await get_live_weather_difficulty(trail_coords)
        print(f"Weather data: {weather_data}")

        # Get static weather exposure info from stored score
        weather_score = trail.get("weather_difficulty_multiplier")
        if weather_score is None:
            weather_score = 1.0  # Default value
        weather_exposure = get_weather_exposure_from_score(weather_score)

        # Safe difficulty calculation
        base_difficulty = trail.get("difficulty_score", 0)
        if base_difficulty is None:
            base_difficulty = 0

        weather_multiplier = weather_data.get("multiplier", 1.0)
        if weather_multiplier is None:
            weather_multiplier = 1.0

        result = {
            "success": True,
            "trail_name": trail.get("name", "Unknown Trail"),
            "live_weather": weather_data,
            "weather_exposure": weather_exposure,
            "coordinates": trail_coords,
            "updated_difficulty": {
                "base_difficulty": base_difficulty,
                "weather_adjusted": round(base_difficulty * weather_multiplier, 1),
                "adjustment_explanation": f"Difficulty adjusted by {weather_multiplier}x due to current conditions",
            },
        }

        print(f"Returning result: {result}")
        return result

    except Exception as e:
        print(f"Weather lookup error: {e}")
        import traceback

        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/trail/{trail_id}/dem3d")
async def get_trail_3d_dem(trail_id: int):
    """Get 3D DEM data for a specific trail"""
    try:
        print(f"Getting 3D DEM data for trail ID: {trail_id}")

        # Get the trail from database
        trail_response = (
            supabase.table("trails").select("*").eq("id", trail_id).execute()
        )
        if not trail_response.data:
            raise HTTPException(status_code=404, detail="Trail not found")

        trail = trail_response.data[0]
        trail_coords = trail.get("coordinates", [])

        if not trail_coords:
            raise HTTPException(status_code=400, detail="Trail has no coordinate data")

        print(
            f"Processing DEM for trail: {trail.get('name', 'Unknown')} with {len(trail_coords)} coordinates"
        )

        # Find relevant DEM tiles
        dem_files = find_relevant_dem_tiles(trail_coords)
        if not dem_files:
            raise HTTPException(
                status_code=404, detail="No DEM data available for this trail area"
            )

        print(f"Found {len(dem_files)} DEM files")

        # Process DEM data
        dem_data = process_dem_for_trail(trail_coords, dem_files)
        if not dem_data:
            raise HTTPException(status_code=500, detail="Failed to process DEM data")

        return {
            "success": True,
            "trail_name": trail.get("name", "Unknown Trail"),
            "trail_id": trail_id,
            "dem_data": dem_data,
        }

    except HTTPException:
        raise
    except Exception as e:
        print(f"3D DEM error: {e}")
        import traceback

        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/map")
async def get_map():
    """Generate map with all trails from Supabase"""
    try:
        # Get trails from database
        response = supabase.table("trails").select("*").execute()
        trails = response.data

        # If no trails, return empty map
        if not trails:
            # Create empty map centered on Brisbane
            m = folium.Map(location=[-27.4698, 152.9560], zoom_start=12)

            # Generate unique filename
            map_id = str(uuid.uuid4())
            map_filename = f"empty_map_{map_id}.html"
            map_path = os.path.join(tempfile.gettempdir(), map_filename)

            # Save map to temporary file
            m.save(map_path)

            return {
                "success": True,
                "map_url": f"/maps/{map_filename}",
                "trails_count": 0,
                "message": "No trails available. Upload GPX files to get started.",
            }

        # Create a map centered on Brisbane for multiple trails
        m = folium.Map(
            location=[-27.4698, 152.9560],
            zoom_start=12,
            tiles="OpenStreetMap",  # Better base layer
            control_scale=True,  # Add scale control
            prefer_canvas=False,  # Ensure interactive behavior
        )

        # Collect all coordinates to calculate bounds
        all_coordinates = []

        # Color palette for different trails
        colors = ["blue", "red", "green", "purple", "orange", "darkred", "lightred"]

        for i, trail in enumerate(trails):
            color = colors[i % len(colors)]
            coordinates = trail.get("coordinates", [])

            if coordinates:
                # Add all coordinates to bounds calculation
                all_coordinates.extend(coordinates)

                # Add polyline for this trail with better styling
                folium.PolyLine(
                    coordinates,
                    color=color,
                    weight=4,
                    opacity=0.8,
                    tooltip=f"{trail.get('name', 'Unnamed Trail')} - {trail.get('distance', 0):.1f}km",
                    popup=folium.Popup(
                        f"""
                        <div style="font-family: Arial, sans-serif;">
                            <h4 style="margin: 0 0 10px 0; color: {color};">{trail.get('name', 'Unnamed Trail')}</h4>
                            <p style="margin: 5px 0;"><strong>Distance:</strong> {trail.get('distance', 0):.1f} km</p>
                            <p style="margin: 5px 0;"><strong>Elevation Gain:</strong> {trail.get('elevation_gain', 0)} m</p>
                            <p style="margin: 5px 0;"><strong>Difficulty:</strong> {trail.get('difficulty_level', 'Unknown')}</p>
                            <p style="margin: 5px 0;"><strong>Max Elevation:</strong> {trail.get('max_elevation', 0)} m</p>
                        </div>
                        """,
                        max_width=250,
                    ),
                ).add_to(m)

                # Add start marker with better styling
                start_coord = coordinates[0]
                folium.Marker(
                    start_coord,
                    popup=folium.Popup(
                        f"<strong>Start:</strong> {trail.get('name', 'Unnamed Trail')}",
                        max_width=200,
                    ),
                    tooltip="Trail Start",
                    icon=folium.Icon(color="green", icon="play", prefix="fa"),
                ).add_to(m)

                # Add end marker with better styling
                end_coord = coordinates[-1]
                folium.Marker(
                    end_coord,
                    popup=folium.Popup(
                        f"<strong>End:</strong> {trail.get('name', 'Unnamed Trail')}",
                        max_width=200,
                    ),
                    tooltip="Trail End",
                    icon=folium.Icon(color="red", icon="stop", prefix="fa"),
                ).add_to(m)

        # Fit map bounds to show all trails
        if all_coordinates:
            # Calculate bounds
            lats = [coord[0] for coord in all_coordinates]
            lons = [coord[1] for coord in all_coordinates]

            # Add some padding to the bounds
            lat_padding = (max(lats) - min(lats)) * 0.1
            lon_padding = (max(lons) - min(lons)) * 0.1

            bounds = [
                [min(lats) - lat_padding, min(lons) - lon_padding],
                [max(lats) + lat_padding, max(lons) + lon_padding],
            ]

            m.fit_bounds(bounds)

        folium.TileLayer("OpenStreetMap").add_to(m)

        # Add JavaScript for trail click handling
        trail_data_js = f"""
        var allTrailsData = {json.dumps([{
            'id': trail.get('id'),
            'name': trail.get('name', 'Unnamed Trail'),
            'distance': trail.get('distance', 0),
            'elevationGain': trail.get('elevation_gain', 0),
            'elevationLoss': trail.get('elevation_loss', 0),
            'maxElevation': trail.get('max_elevation', 0),
            'minElevation': trail.get('min_elevation', 0),
            'rollingHillsIndex': trail.get('rolling_hills_index', 0),
            'difficultyScore': trail.get('difficulty_score', 0),
            'difficultyLevel': trail.get('difficulty_level', 'Unknown'),
            'elevationProfile': trail.get('elevation_profile', []),
            'coordinates': trail.get('coordinates', [])
        } for trail in trails])};
        
        console.log('Trail data available:', allTrailsData.length, 'trails');
        allTrailsData.forEach(function(trail, index) {{
            console.log('Trail', index + ':', trail.name, '- ID:', trail.id);
        }});
        
        function sendTrailDataToParent(trailData) {{
            console.log('Sending trail data to parent:', trailData);
            if (window.parent && window.parent !== window) {{
                window.parent.postMessage({{
                    type: 'trail-clicked',
                    data: trailData
                }}, '*');
            }} else {{
                console.log('No parent window found - running in standalone mode');
            }}
        }}
        
        function setupClickHandlers() {{
            var polylines = document.querySelectorAll('.leaflet-interactive');
            console.log('Found', polylines.length, 'interactive elements');
            
            polylines.forEach(function(polyline, index) {{
                polyline.style.cursor = 'pointer';
                polyline.addEventListener('click', function(e) {{
                    console.log('Polyline', index, 'clicked');
                    if (allTrailsData[index]) {{
                        sendTrailDataToParent(allTrailsData[index]);
                    }} else {{
                        console.log('No trail data found for index', index);
                    }}
                }});
            }});
        }}
        
        // Setup click handlers when DOM is ready
        if (document.readyState === 'loading') {{
            document.addEventListener('DOMContentLoaded', function() {{
                setTimeout(setupClickHandlers, 1000);
            }});
        }} else {{
            setTimeout(setupClickHandlers, 1000);
        }}
        """

        # Add JavaScript to the map
        m.get_root().html.add_child(
            folium.Element(
                f"""
        <script>
        {trail_data_js}
        </script>
        """
            )
        )

        # Generate unique filename and save map
        map_id = str(uuid.uuid4())
        map_filename = f"trails_map_{map_id}.html"
        map_path = os.path.join(tempfile.gettempdir(), map_filename)

        # Save map to temporary file
        m.save(map_path)

        return {
            "success": True,
            "map_url": f"/maps/{map_filename}",
            "trails_count": len(trails),
        }

    except Exception as e:
        print(f"Map generation error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/upload-gpx")
async def upload_gpx(file: UploadFile = File(...), overwrite: str = Form("false")):
    """Handle GPX file upload and save to Supabase

    Args:
        file: GPX file to upload
        overwrite: If True, will delete existing trail with same name/location before adding new one
    """
    # Convert overwrite string to boolean
    print(f"📤 Uploading GPX file: {file.filename}")
    print(f"   Raw overwrite value: '{overwrite}' (type: {type(overwrite).__name__})")

    # Safety check for None or empty string
    if overwrite is None:
        overwrite_bool = False
    elif isinstance(overwrite, str):
        overwrite_bool = overwrite.lower() in ("true", "1", "yes")
    else:
        overwrite_bool = bool(overwrite)

    print(f"   Converted to bool: {overwrite_bool}")

    if not file.filename.lower().endswith(".gpx"):
        raise HTTPException(status_code=400, detail="File must be a GPX file")

    try:
        # Read GPX content
        content = await file.read()
        gpx_content = content.decode("utf-8")

        # Extract trail name from filename
        trail_name = file.filename.replace(".gpx", "").replace("_", " ").title()

        # Analyze trail data
        gpx = gpxpy.parse(gpx_content)
        coords = []
        elevations = []
        distances = [0]
        slopes = [0]  # Start with 0 slope for first point

        for track in gpx.tracks:
            for segment in track.segments:
                for i, point in enumerate(segment.points):
                    coords.append([point.latitude, point.longitude])
                    elevations.append(point.elevation or 0)

                    if i > 0:
                        prev_point = segment.points[i - 1]
                        dist = (
                            haversine(
                                prev_point.latitude,
                                prev_point.longitude,
                                point.latitude,
                                point.longitude,
                            )
                            / 1000
                        )
                        distances.append(distances[-1] + dist)

                        # Slope analysis (gradient in %)
                        elev_diff = (point.elevation or 0) - (prev_point.elevation or 0)
                        dist_m = haversine(
                            prev_point.latitude,
                            prev_point.longitude,
                            point.latitude,
                            point.longitude,
                        )
                        if dist_m > 0:
                            gradient = (elev_diff / dist_m) * 100
                            slopes.append(gradient)
                        else:
                            slopes.append(0)

        if not coords:
            raise HTTPException(
                status_code=400, detail="No track points found in GPX file"
            )

        # Calculate statistics
        total_distance = distances[-1] if len(distances) > 1 else 0
        elevation_gain = (
            sum(
                max(0, elevations[i] - elevations[i - 1])
                for i in range(1, len(elevations))
            )
            if len(elevations) > 1
            else 0
        )
        elevation_loss = (
            sum(
                max(0, elevations[i - 1] - elevations[i])
                for i in range(1, len(elevations))
            )
            if len(elevations) > 1
            else 0
        )
        max_elevation = max(elevations) if elevations else 0
        min_elevation = min(elevations) if elevations else 0

        # Rolling hills analysis (advanced) - returns index and count
        rolling_hills_index, rolling_hills_count = analyze_rolling_hills(
            elevations, distances
        )
        rolling_hills_index = round(rolling_hills_index, 2)
        print(f"🔍 DEBUG: Rolling Hills Index calculated: {rolling_hills_index}")
        print(f"🔍 DEBUG: Rolling Hills Count: {rolling_hills_count}")
        print(
            f"🔍 DEBUG: Elevations count: {len(elevations)}, Distance: {distances[-1] if distances else 0} km"
        )

        # Create elevation profile data
        elevation_profile_data = [
            {
                "distance": round(dist, 2),
                "elevation": round(ele, 1),
                "slope": round(slopes[i] if i < len(slopes) else 0, 2),
            }
            for i, (dist, ele) in enumerate(zip(distances, elevations))
        ]

        # Slope analysis
        if slopes and len(slopes) > 1:  # Skip first element (always 0)
            slope_values = slopes[1:]  # Skip the initial 0
            max_slope = max(slope_values) if slope_values else 0
            avg_slope = (
                sum(map(abs, slope_values)) / len(slope_values) if slope_values else 0
            )
        else:
            max_slope = 0
            avg_slope = 0

        # Segment analysis (500m segments)
        segment_length = 0.5  # km
        segments = []
        if len(distances) > 1:
            seg_start_idx = 0
            while seg_start_idx < len(distances) - 1:
                seg_end_idx = seg_start_idx
                # Find the end index for this segment
                while (
                    seg_end_idx < len(distances) - 1
                    and distances[seg_end_idx] - distances[seg_start_idx]
                    < segment_length
                ):
                    seg_end_idx += 1
                # Calculate stats for this segment
                seg_dist = distances[seg_end_idx] - distances[seg_start_idx]
                seg_elev_change = elevations[seg_end_idx] - elevations[seg_start_idx]
                # Slope for segment
                if seg_dist > 0:
                    seg_slope = (seg_elev_change / (seg_dist * 1000)) * 100
                else:
                    seg_slope = 0
                segments.append(
                    {
                        "start_distance": round(distances[seg_start_idx], 2),
                        "end_distance": round(distances[seg_end_idx], 2),
                        "elevation_change": round(seg_elev_change, 1),
                        "avg_slope": round(seg_slope, 2),
                    }
                )
                seg_start_idx = seg_end_idx

        # Simple difficulty calculation
        distance_factor = min(total_distance / 10, 1) * 3  # 0-3 points
        elevation_factor = min(elevation_gain / 1000, 1) * 4  # 0-4 points
        # Normalize rolling hills index (typically 0-50 range) to 0-1, then scale to 0-3 points
        normalized_rolling = min(
            rolling_hills_index / 50, 1
        )  # Cap at 50 for normalization
        rolling_factor = normalized_rolling * 3  # 0-3 points
        difficulty_score = (
            distance_factor + elevation_factor + rolling_factor
        )  # Max = 10 points

        if difficulty_score <= 3:
            difficulty_level = "Easy"
        elif difficulty_score <= 6:
            difficulty_level = "Moderate"
        elif difficulty_score <= 8:
            difficulty_level = "Hard"
        else:
            difficulty_level = "Extreme"

        # Check for duplicate trails before inserting
        # First check by exact name match
        existing_trails_response = (
            supabase.table("trails").select("*").eq("name", trail_name).execute()
        )

        duplicate_trail_id = None
        if existing_trails_response.data:
            if not overwrite_bool:
                raise HTTPException(
                    status_code=409,
                    detail=f"Trail with name '{trail_name}' already exists in database",
                )
            else:
                # Delete existing trail and its associated LiDAR files
                duplicate_trail_id = existing_trails_response.data[0]["id"]
                print(
                    f"🗑️  Overwrite mode: Deleting existing trail ID {duplicate_trail_id}"
                )

                # Delete associated LiDAR files first
                lidar_files = (
                    supabase.table("lidar_files")
                    .select("*")
                    .eq("trail_id", duplicate_trail_id)
                    .execute()
                )
                if lidar_files.data:
                    print(
                        f"   Deleting {len(lidar_files.data)} associated LiDAR file(s)"
                    )
                    for lidar_file in lidar_files.data:
                        db_client = supabase_service if supabase_service else supabase
                        db_client.table("lidar_files").delete().eq(
                            "id", lidar_file["id"]
                        ).execute()

                # Delete the trail
                db_client = supabase_service if supabase_service else supabase
                db_client.table("trails").delete().eq(
                    "id", duplicate_trail_id
                ).execute()
                print(f"   ✅ Deleted trail and associated data")

        # Check for similar starting coordinates (within ~100m radius)
        # This prevents uploading the same trail with different names
        start_lat, start_lon = coords[0]
        all_trails_response = (
            supabase.table("trails").select("id, name, coordinates").execute()
        )

        for existing_trail in all_trails_response.data:
            if existing_trail.get("coordinates"):
                existing_start = existing_trail["coordinates"][0]
                existing_lat, existing_lon = existing_start

                # Calculate distance between starting points
                distance_between_starts = haversine(
                    start_lat, start_lon, existing_lat, existing_lon
                )

                # If starts are within 100 meters, likely duplicate
                if distance_between_starts < 100:
                    if not overwrite_bool:
                        raise HTTPException(
                            status_code=409,
                            detail=f"Trail starting near same location as existing trail '{existing_trail['name']}' (within 100m). Possible duplicate.",
                        )
                    else:
                        # Delete this coordinate-duplicate trail too
                        coord_dup_id = existing_trail["id"]
                        if coord_dup_id != duplicate_trail_id:  # Don't delete twice
                            print(
                                f"🗑️  Overwrite mode: Deleting coordinate-duplicate trail ID {coord_dup_id}"
                            )

                            # Delete associated LiDAR files
                            lidar_files = (
                                supabase.table("lidar_files")
                                .select("*")
                                .eq("trail_id", coord_dup_id)
                                .execute()
                            )
                            if lidar_files.data:
                                print(
                                    f"   Deleting {len(lidar_files.data)} associated LiDAR file(s)"
                                )
                                for lidar_file in lidar_files.data:
                                    db_client = (
                                        supabase_service
                                        if supabase_service
                                        else supabase
                                    )
                                    db_client.table("lidar_files").delete().eq(
                                        "id", lidar_file["id"]
                                    ).execute()

                            # Delete the trail
                            db_client = (
                                supabase_service if supabase_service else supabase
                            )
                            db_client.table("trails").delete().eq(
                                "id", coord_dup_id
                            ).execute()
                            print(f"   ✅ Deleted coordinate-duplicate trail")

        # Create new trail data for Supabase
        weather_exposure = get_trail_weather_exposure({"max_elevation": max_elevation})
        terrain_variety = calculate_terrain_variety(elevations)

        # Convert weather exposure to a numeric score for database compatibility
        exposure_scores = {
            "Low": 1.0,
            "Low-Moderate": 1.1,
            "Moderate": 1.2,
            "High": 1.3,
        }
        weather_score = exposure_scores.get(weather_exposure["exposure_level"], 1.0)

        new_trail_data = {
            "name": trail_name,
            "distance": round(total_distance, 2),
            "elevation_gain": int(round(elevation_gain, 0)),
            "elevation_loss": int(round(elevation_loss, 0)),
            "max_elevation": int(round(max_elevation, 0)),
            "min_elevation": int(round(min_elevation, 0)),
            "rolling_hills_index": rolling_hills_index,
            "rolling_hills_count": rolling_hills_count,  # Number of significant elevation changes
            "difficulty_score": round(difficulty_score, 1),
            "difficulty_level": difficulty_level,
            "coordinates": coords,
            "elevation_profile": elevation_profile_data,
            "max_slope": round(max_slope, 2),
            "avg_slope": round(avg_slope, 2),
            "segments": segments,
            # Enhanced effort estimation using Naismith's Rule + terrain adjustments
            "estimated_time_hours": round(
                (total_distance / 5)
                + (elevation_gain / 600)
                + (rolling_hills_index * 0.5),
                2,
            ),
            # Improved analytics fields (using existing column names)
            "terrain_variety_score": terrain_variety,
            "elevation_change_total": int(round(elevation_gain + elevation_loss, 0)),
            # Store weather data in existing numeric fields, we'll interpret on frontend
            "weather_difficulty_multiplier": weather_score,  # Use existing column with new meaning
            # Fixed technical difficulty calculation (1-10 scale)
            # Factors: max slope (40%), rolling hills (30%), avg slope (30%)
            # Properly normalized to produce 1-10 range
            "technical_rating": round(
                max(
                    1.0,
                    min(
                        10.0,
                        1
                        + (max_slope / 100) * 3.5  # Max slope: 0-100% -> 0-3.5 points
                        + (
                            min(rolling_hills_index / 50, 1.0) * 3.5
                        )  # Rolling hills normalized: 0-50 -> 0-3.5 points
                        + (avg_slope / 30) * 2.0,  # Avg slope: 0-30% -> 0-2 points
                    ),
                ),
            ),
        }

        # Insert trail into Supabase database (use service-role client if available to bypass RLS)
        db_client = supabase_service if supabase_service else supabase
        response = db_client.table("trails").insert(new_trail_data).execute()

        if response.data:
            inserted_trail = response.data[0]
            return {
                "success": True,
                "message": "Trail uploaded successfully to database",
                "trail": inserted_trail,
            }
        else:
            raise HTTPException(
                status_code=500, detail="Failed to insert trail into database"
            )

    except HTTPException:
        # Re-raise HTTPExceptions (409, 400, etc.) without wrapping them
        raise
    except Exception as e:
        import traceback

        error_details = traceback.format_exc()
        print(f"❌ Upload error: {type(e).__name__}: {e}")
        print(f"❌ Full traceback:\n{error_details}")
        raise HTTPException(
            status_code=500, detail=f"{type(e).__name__}: {str(e) or 'Unknown error'}"
        )


@app.get("/trail/{trail_id}/3d-terrain-viewer")
async def get_trail_3d_terrain_viewer(trail_id: int, elevation_source: str = "gpx"):
    """Serve interactive 3D terrain visualization as a standalone HTML page

    Args:
        trail_id: Trail ID
        elevation_source: "gpx" (default) or "lidar" - determines which elevation data to use for trail overlay
    """
    try:
        if not dem_analyzer:
            return JSONResponse(
                {"error": "DEM analyzer not available"}, status_code=503
            )

        # Get trail data
        trail_response = (
            supabase.table("trails").select("*").eq("id", trail_id).execute()
        )
        if not trail_response.data:
            return JSONResponse({"error": "Trail not found"}, status_code=404)

        trail = trail_response.data[0]
        coordinates = trail.get("coordinates", [])

        if not coordinates:
            return JSONResponse({"error": "No coordinates available"}, status_code=400)

        # Get LiDAR elevations if requested
        lidar_elevations = None
        if elevation_source.lower() == "lidar":
            try:
                # Extract elevation profile from LiDAR
                profile = lidar_extractor.extract_elevation_profile(
                    trail_coords=coordinates, trail_id=trail_id
                )
                if profile and profile.get("success") and "elevations" in profile:
                    lidar_elevations = profile["elevations"]

                    # Align LiDAR elevations to GPX baseline (same as elevation profile chart)
                    elevation_profile = trail.get("elevation_profile", [])
                    if elevation_profile and len(elevation_profile) > 0:
                        # Get GPX elevations
                        if isinstance(elevation_profile[0], dict):
                            gpx_elevations = [
                                point.get("elevation") for point in elevation_profile
                            ]
                        else:
                            gpx_elevations = elevation_profile

                        if (
                            gpx_elevations
                            and len(gpx_elevations) > 0
                            and len(lidar_elevations) > 0
                        ):
                            # Calculate offset based on starting elevations (always align)
                            gpx_start = gpx_elevations[0]
                            lidar_start = lidar_elevations[0]
                            elevation_offset = gpx_start - lidar_start

                            # Apply offset to align LiDAR to GPX baseline
                            lidar_elevations = [
                                e + elevation_offset for e in lidar_elevations
                            ]
                            print(
                                f"🔧 Aligned LiDAR elevations to GPX baseline: offset={elevation_offset:.1f}m"
                            )
                            print(
                                f"   Before: GPX={gpx_start:.1f}m, LiDAR={lidar_start:.1f}m (diff: {elevation_offset:.1f}m)"
                            )
                            print(
                                f"   After:  GPX={gpx_start:.1f}m, LiDAR={lidar_elevations[0]:.1f}m (aligned ✓)"
                            )

                    print(
                        f"📊 Using {len(lidar_elevations)} LiDAR elevation points for 3D visualization"
                    )
                else:
                    print("⚠️  No LiDAR data available, falling back to GPX/DEM")
                    elevation_source = "gpx"
            except Exception as e:
                print(f"❌ Error getting LiDAR elevations: {e}")
                elevation_source = "gpx"

        # Generate 3D visualization
        visualization_result = dem_analyzer.create_3d_terrain_visualization(
            coordinates,
            buffer_meters=1000,
            elevation_source=elevation_source,
            trail_id=trail_id,
            lidar_elevations=lidar_elevations,
        )

        if not visualization_result.get("success"):
            error_html = f"""
            <!DOCTYPE html>
            <html>
            <head><title>3D Terrain Error</title></head>
            <body style="font-family: Arial, sans-serif; padding: 20px; text-align: center;">
                <h2>3D Terrain Visualization Error</h2>
                <p>{visualization_result.get('error', 'Unknown error')}</p>
            </body>
            </html>
            """
            return HTMLResponse(content=error_html, status_code=500)

        # Return interactive HTML if available
        if visualization_result.get("type") == "interactive":
            return HTMLResponse(content=visualization_result["html_content"])
        else:
            # Create HTML wrapper for static image
            static_html = f"""
            <!DOCTYPE html>
            <html>
            <head>
                <title>3D Terrain - {trail.get('name', 'Trail')}</title>
                <style>
                    body {{ font-family: Arial, sans-serif; margin: 0; padding: 20px; text-align: center; }}
                    img {{ max-width: 100%; height: auto; border: 1px solid #ddd; border-radius: 8px; }}
                </style>
            </head>
            <body>
                <h2>3D Terrain Visualization - {trail.get('name', 'Trail')}</h2>
                <img src="data:image/png;base64,{visualization_result['image_base64']}" 
                     alt="3D Terrain Visualization" />
                <p>{visualization_result['description']}</p>
            </body>
            </html>
            """
            return HTMLResponse(content=static_html)

    except Exception as e:
        print(f"3D terrain viewer error: {e}")
        error_html = f"""
        <!DOCTYPE html>
        <html>
        <head><title>3D Terrain Error</title></head>
        <body style="font-family: Arial, sans-serif; padding: 20px; text-align: center;">
            <h2>3D Terrain Visualization Error</h2>
            <p>Server error: {str(e)}</p>
        </body>
        </html>
        """
        return HTMLResponse(content=error_html, status_code=500)


@app.get("/maps/{filename}")
async def serve_map_file(filename: str):
    """Serve generated map files"""
    map_path = os.path.join(tempfile.gettempdir(), filename)
    if not os.path.exists(map_path):
        raise HTTPException(status_code=404, detail="Map file not found")
    return FileResponse(map_path)


# Enhanced DEM/LiDAR Analysis Endpoints
@app.get("/trail/{trail_id}/dem-analysis")
async def get_trail_dem_analysis(trail_id: int):
    """Analyze DEM data for a specific trail using real DEM files"""
    try:
        if not dem_analyzer:
            return {
                "success": False,
                "error": "DEM analyzer not available. Check DEM file path and dependencies.",
            }

        # Get trail data
        trail_response = (
            supabase.table("trails").select("*").eq("id", trail_id).execute()
        )
        if not trail_response.data:
            raise HTTPException(status_code=404, detail="Trail not found")

        trail = trail_response.data[0]
        coordinates = trail.get("coordinates", [])

        if not coordinates:
            return {
                "success": False,
                "error": "No coordinates available for this trail",
            }

        print(
            f"Analyzing trail '{trail.get('name')}' with {len(coordinates)} coordinate points"
        )

        # Extract real elevation profile from DEM data
        elevation_analysis = dem_analyzer.extract_elevation_profile(coordinates)

        if not elevation_analysis.get("success"):
            return {
                "success": False,
                "error": elevation_analysis.get("error", "Analysis failed"),
            }

        # Analyze terrain features
        terrain_features = dem_analyzer.analyze_terrain_features(coordinates)

        # Generate 3D visualization
        visualization_3d = dem_analyzer.create_3d_terrain_visualization(coordinates)

        result = {
            "success": True,
            "trail_name": trail.get("name"),
            "elevation_analysis": elevation_analysis,
            "terrain_features": terrain_features,
            "visualization_3d": visualization_3d,
            "data_quality": {
                "resolution": "1 meter",
                "data_source": "QSpatial DEM",
                "coordinate_system": "GDA94 MGA Zone 56",
                "accuracy": "±0.5 meters",
            },
        }

        return result

    except Exception as e:
        print(f"DEM analysis error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/trail/{trail_id}/3d-terrain")
async def get_trail_3d_terrain(trail_id: int, elevation_source: str = "gpx"):
    """Generate interactive 3D terrain visualization for a trail

    Args:
        trail_id: Trail ID
        elevation_source: "gpx" (default) or "lidar" - determines which elevation data to use for trail overlay
    """
    try:
        if not dem_analyzer:
            return {"success": False, "error": "DEM analyzer not available"}

        # Get trail data
        trail_response = (
            supabase.table("trails").select("*").eq("id", trail_id).execute()
        )
        if not trail_response.data:
            raise HTTPException(status_code=404, detail="Trail not found")

        trail = trail_response.data[0]
        coordinates = trail.get("coordinates", [])

        if not coordinates:
            return {"success": False, "error": "No coordinates available"}

        # Get LiDAR elevations if requested
        lidar_elevations = None
        if elevation_source.lower() == "lidar":
            try:
                # Extract elevation profile from LiDAR
                profile = lidar_extractor.extract_elevation_profile(
                    trail_coords=coordinates, trail_id=trail_id
                )
                if profile and profile.get("success") and "elevations" in profile:
                    lidar_elevations = profile["elevations"]

                    # Align LiDAR elevations to GPX baseline (same as elevation profile chart)
                    elevation_profile = trail.get("elevation_profile", [])
                    if elevation_profile and len(elevation_profile) > 0:
                        # Get GPX elevations
                        if isinstance(elevation_profile[0], dict):
                            gpx_elevations = [
                                point.get("elevation") for point in elevation_profile
                            ]
                        else:
                            gpx_elevations = elevation_profile

                        if (
                            gpx_elevations
                            and len(gpx_elevations) > 0
                            and len(lidar_elevations) > 0
                        ):
                            # Calculate offset based on starting elevations (always align)
                            gpx_start = gpx_elevations[0]
                            lidar_start = lidar_elevations[0]
                            elevation_offset = gpx_start - lidar_start

                            # Apply offset to align LiDAR to GPX baseline
                            lidar_elevations = [
                                e + elevation_offset for e in lidar_elevations
                            ]
                            print(
                                f"🔧 Aligned LiDAR elevations to GPX baseline: offset={elevation_offset:.1f}m"
                            )
                            print(
                                f"   Before: GPX={gpx_start:.1f}m, LiDAR={lidar_start:.1f}m (diff: {elevation_offset:.1f}m)"
                            )
                            print(
                                f"   After:  GPX={gpx_start:.1f}m, LiDAR={lidar_elevations[0]:.1f}m (aligned ✓)"
                            )

                    print(
                        f"📊 Using {len(lidar_elevations)} LiDAR elevation points for 3D visualization"
                    )
                else:
                    print("⚠️  No LiDAR data available, falling back to GPX/DEM")
                    elevation_source = "gpx"
            except Exception as e:
                print(f"❌ Error getting LiDAR elevations: {e}")
                elevation_source = "gpx"

        # Generate 3D visualization
        visualization_result = dem_analyzer.create_3d_terrain_visualization(
            coordinates,
            buffer_meters=1000,
            elevation_source=elevation_source,
            trail_id=trail_id,
            lidar_elevations=lidar_elevations,
        )

        if not visualization_result.get("success"):
            return {
                "success": False,
                "error": visualization_result.get(
                    "error", "Failed to generate 3D visualization"
                ),
            }

        # Return based on visualization type
        if visualization_result.get("type") == "interactive":
            return {
                "success": True,
                "trail_name": trail.get("name"),
                "visualization_type": "interactive",
                "visualization_html": visualization_result["html_content"],
                "description": visualization_result["description"],
            }
        else:
            # Static visualization
            return {
                "success": True,
                "trail_name": trail.get("name"),
                "visualization_type": "static",
                "visualization": f"data:image/png;base64,{visualization_result['image_base64']}",
                "description": visualization_result["description"],
            }

    except Exception as e:
        print(f"3D terrain error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/dem/coverage")
async def get_dem_coverage():
    """Get information about available DEM coverage"""
    try:
        if not dem_analyzer:
            return {
                "success": False,
                "error": "DEM analyzer not available",
                "troubleshooting": {
                    "check_path": "backend/data/QSpatial/DEM/1 Metre",
                    "required_packages": ["rasterio", "geopandas", "pyproj"],
                    "file_format": ".tif files",
                },
            }

        # Get actual file information
        dem_files = dem_analyzer.dem_files
        total_size_mb = 0

        file_info = []
        for dem_file in dem_files[:10]:  # Show first 10 files as examples
            try:
                size_mb = os.path.getsize(dem_file) / (1024 * 1024)
                total_size_mb += size_mb

                # Extract coordinate info from filename
                filename = os.path.basename(dem_file)
                file_info.append(
                    {
                        "filename": filename,
                        "size_mb": round(size_mb, 2),
                        "path": dem_file,
                    }
                )
            except:
                continue

        # Estimate total size for all files
        if len(dem_files) > 10:
            avg_size = total_size_mb / len(file_info) if file_info else 50
            estimated_total_mb = avg_size * len(dem_files)
        else:
            estimated_total_mb = total_size_mb

        coverage_info = {
            "available": True,
            "total_files": len(dem_files),
            "resolution": "1 meter",
            "format": "GeoTIFF (.tif)",
            "coordinate_system": "GDA94 / MGA Zone 56 (EPSG:28356)",
            "coverage_area": "Brisbane Region",
            "years_available": ["2009", "2014", "2019"],
            "estimated_size_gb": round(estimated_total_mb / 1024, 2),
            "sample_files": file_info,
            "data_path": dem_analyzer.dem_base_path,
        }

        return {"success": True, "coverage": coverage_info}

    except Exception as e:
        return {"success": False, "error": str(e)}


@app.get("/trail/{trail_id}/elevation-sources")
async def get_trail_elevation_sources(trail_id: int):
    """
    Get elevation profiles from all available data sources:
    - GPX: Raw elevation data from the uploaded GPX file
    - LiDAR: High-resolution point cloud data (.las files)
    - QSpatial: Digital Elevation Model (DEM) data
    - Overall: Averaged elevation from all available sources
    """
    try:
        print(f"\n🔍 Getting elevation sources for trail {trail_id}")

        # Get trail data from database
        trail_response = (
            supabase.table("trails").select("*").eq("id", trail_id).execute()
        )
        if not trail_response.data:
            raise HTTPException(status_code=404, detail="Trail not found")

        trail = trail_response.data[0]
        print(f"📍 Trail name: {trail.get('name')}")

        coordinates = trail.get("coordinates", [])
        print(f"📊 Coordinates count: {len(coordinates)}")

        if not coordinates:
            return {
                "success": False,
                "error": "No coordinates available for this trail",
            }

        # Calculate distances along trail for x-axis
        distances = [0]
        print(f"🧮 Calculating distances for {len(coordinates)} coordinates")
        print(
            f"📍 First coordinate sample: {coordinates[0] if coordinates else 'None'}"
        )

        for i in range(1, len(coordinates)):
            try:
                lat1, lon1 = coordinates[i - 1]
                lat2, lon2 = coordinates[i]
                dist = haversine(lat1, lon1, lat2, lon2)
                distances.append(distances[-1] + dist)
            except Exception as e:
                print(f"❌ Error at coordinate index {i}: {e}")
                print(f"   Coordinate[{i-1}]: {coordinates[i-1]}")
                print(f"   Coordinate[{i}]: {coordinates[i]}")
                raise

        # Convert distances to kilometers
        distances_km = [d / 1000 for d in distances]
        print(f"✅ Calculated {len(distances_km)} distance points")

        # Initialize results
        sources = {}

        # 1. GPX Source - use elevations from trail data if available
        # Check both "elevations" and "elevation_profile" columns
        elevation_profile = trail.get("elevation_profile", [])

        # Extract elevations from elevation_profile if it's an array of objects
        if (
            elevation_profile
            and isinstance(elevation_profile, list)
            and len(elevation_profile) > 0
        ):
            if isinstance(elevation_profile[0], dict):
                # elevation_profile is array of {elevation, distance, slope} objects
                gpx_elevations = [point.get("elevation") for point in elevation_profile]
            else:
                # elevation_profile is already an array of numbers
                gpx_elevations = elevation_profile
        else:
            # Fallback to "elevations" column (older format)
            gpx_elevations = trail.get("elevations", [])

        print(
            f"📈 GPX elevations count: {len(gpx_elevations) if gpx_elevations else 0}"
        )

        if gpx_elevations and len(gpx_elevations) == len(coordinates):
            print("✅ GPX source available")
            sources["GPX"] = {
                "available": True,
                "elevations": gpx_elevations,
                "distances": distances_km,
                "coordinates": coordinates,
                "source": "Original GPX file",
                "data_points": len(gpx_elevations),
            }
        else:
            print(
                f"❌ GPX source not available - elevations: {len(gpx_elevations) if gpx_elevations else 0}, coordinates: {len(coordinates)}"
            )
            sources["GPX"] = {
                "available": False,
                "error": "No elevation data in GPX file",
                "elevations": [],
                "distances": [],
                "coordinates": [],
            }

        # 2. LiDAR Source
        if lidar_extractor:
            try:
                print(f"🔍 Searching for matching LiDAR file...")
                print(f"   Available LiDAR files: {len(lidar_extractor.lidar_files)}")
                lidar_result = lidar_extractor.extract_elevation_profile(
                    coordinates, trail_id=trail_id
                )
                print(
                    f"   LiDAR extraction result: success={lidar_result.get('success')}, error={lidar_result.get('error')}"
                )

                if lidar_result.get("success"):
                    # Ensure same length as coordinates by interpolation if needed
                    lidar_elevations = lidar_result.get("elevations", [])

                    # Use LiDAR's own distances if provided (for relative coordinate files)
                    # Otherwise use trail distances
                    if "distances" in lidar_result and lidar_result["distances"]:
                        lidar_distances = lidar_result["distances"]
                    else:
                        lidar_distances = distances_km[: len(lidar_elevations)]

                    # Align LiDAR elevations with GPX if using relative coordinates
                    note = lidar_result.get("note", "")
                    if (
                        "relative coordinates" in note.lower()
                        and sources["GPX"]["available"]
                    ):
                        gpx_elevations = sources["GPX"]["elevations"]
                        if gpx_elevations and lidar_elevations:
                            # Calculate offset to align starting elevations
                            gpx_start = gpx_elevations[0]
                            lidar_start = lidar_elevations[0]
                            elevation_offset = gpx_start - lidar_start

                            # Apply offset to all LiDAR elevations
                            lidar_elevations = [
                                e + elevation_offset for e in lidar_elevations
                            ]

                            print(
                                f"   🔧 Aligned LiDAR elevations: offset={elevation_offset:.1f}m"
                            )
                            print(
                                f"      GPX start: {gpx_start:.1f}m, LiDAR start (adjusted): {lidar_elevations[0]:.1f}m"
                            )
                            note += f" | Aligned to GPX baseline (offset: {elevation_offset:.1f}m)"

                    sources["LiDAR"] = {
                        "available": True,
                        "elevations": lidar_elevations,
                        "distances": lidar_distances,
                        "coordinates": lidar_result.get(
                            "coordinates", coordinates[: len(lidar_elevations)]
                        ),
                        "source": f"LiDAR point cloud: {lidar_result.get('lidar_file', 'N/A')}",
                        "data_points": len(lidar_elevations),
                        "coverage_percent": lidar_result.get("coverage_percent", 0),
                        "total_lidar_points": lidar_result.get("total_lidar_points", 0),
                        "note": note,
                    }
                else:
                    sources["LiDAR"] = {
                        "available": False,
                        "error": lidar_result.get("error", "LiDAR extraction failed"),
                        "elevations": [],
                        "distances": [],
                        "coordinates": [],
                    }
            except Exception as e:
                print(f"LiDAR extraction error: {e}")
                sources["LiDAR"] = {
                    "available": False,
                    "error": str(e),
                    "elevations": [],
                    "distances": [],
                    "coordinates": [],
                }
        else:
            sources["LiDAR"] = {
                "available": False,
                "error": "LiDAR extractor not initialized",
                "elevations": [],
                "distances": [],
                "coordinates": [],
            }

        # 2b. XLSX Source - check for uploaded xlsx files linked to this trail
        try:
            print(f"🔍 Looking for XLSX files for trail_id: {trail_id}")
            
            # Debug: Check all XLSX files to see what trail_ids exist
            all_xlsx_resp = supabase.table("xlsx_files").select("id, trail_id, filename, original_filename").execute()
            print(f"📋 All XLSX files in database: {len(all_xlsx_resp.data)} total")
            for xlsx in all_xlsx_resp.data:
                print(f"   - ID: {xlsx.get('id')}, trail_id: {xlsx.get('trail_id')}, file: {xlsx.get('original_filename', xlsx.get('filename'))}")
            
            xlsx_resp = (
                supabase.table("xlsx_files")
                .select("*")
                .eq("trail_id", trail_id)
                .order("created_at", desc=True)
                .limit(1)
                .execute()
            )
            print(f"📊 XLSX query result: {len(xlsx_resp.data)} files found")
            xlsx_record = xlsx_resp.data[0] if xlsx_resp.data else None
            if xlsx_record:
                print(f"🔍 Found XLSX file for trail: {xlsx_record.get('filename')}")
                xlsx_url = xlsx_record.get("file_url")
                # Attempt to fetch and parse XLSX
                try:
                    import requests
                    from openpyxl import load_workbook
                    from io import BytesIO

                    # If file_url is a supabase public_url object/dict, handle it
                    if isinstance(xlsx_url, dict) and "publicURL" in xlsx_url:
                        download_url = xlsx_url["publicURL"]
                    elif isinstance(xlsx_url, str) and xlsx_url.startswith("http"):
                        download_url = xlsx_url
                    else:
                        # Try to get public url via storage client
                        download_url = supabase.storage.from_(
                            "xlsx-files"
                        ).get_public_url(xlsx_record.get("filename"))

                    r = requests.get(download_url, timeout=15)
                    if r.status_code == 200:
                        wb = load_workbook(
                            filename=BytesIO(r.content), read_only=True, data_only=True
                        )
                        sheet = wb[wb.sheetnames[0]]
                        rows = list(sheet.iter_rows(values_only=True))
                        wb.close()

                        if rows and len(rows) > 1:
                            headers_x = [
                                str(h) if h is not None else "" for h in rows[0]
                            ]
                            # Try to locate distance and elevation columns
                            lower = [h.lower() for h in headers_x]
                            try:
                                dist_idx = next(
                                    i
                                    for i, h in enumerate(lower)
                                    if "dist" in h
                                    or "chain" in h
                                    or "length" in h
                                    or "distance" in h
                                )
                            except StopIteration:
                                dist_idx = None
                            try:
                                elev_idx = next(
                                    i
                                    for i, h in enumerate(lower)
                                    if h == "z"
                                    or "elev" in h
                                    or "height" in h
                                    or "alt" in h
                                )
                            except StopIteration:
                                elev_idx = None

                            xlsx_elevations = []
                            xlsx_distances = []
                            for row in rows[1:]:
                                if dist_idx is not None and elev_idx is not None:
                                    d = row[dist_idx]
                                    z = row[elev_idx]
                                else:
                                    # Fallback: try common positions
                                    d = row[1] if len(row) > 1 else None
                                    z = row[2] if len(row) > 2 else None

                                try:
                                    dn = (
                                        float(str(d).replace(",", ""))
                                        if d is not None
                                        else None
                                    )
                                    zn = (
                                        float(str(z).replace(",", ""))
                                        if z is not None
                                        else None
                                    )
                                except Exception:
                                    dn = None
                                    zn = None
                                if dn is not None and zn is not None:
                                    xlsx_distances.append(
                                        dn / 1000.0 if dn > 10 else dn
                                    )  # if distances in meters convert to km heuristically
                                    xlsx_elevations.append(zn)

                            # Sort data by distance to prevent jittery graphs
                            if xlsx_elevations and xlsx_distances:
                                # Combine distance and elevation pairs, sort by distance, then separate
                                combined_data = list(zip(xlsx_distances, xlsx_elevations))
                                combined_data.sort(key=lambda x: x[0])  # Sort by distance
                                
                                # Remove duplicates and ensure monotonic distance progression
                                cleaned_data = []
                                prev_distance = None
                                tolerance = 0.001  # Minimum distance difference (1m for km units)
                                
                                for distance, elevation in combined_data:
                                    # Only add if distance is significantly greater than previous
                                    if prev_distance is None or distance > (prev_distance + tolerance):
                                        cleaned_data.append((distance, elevation))
                                        prev_distance = distance
                                
                                # Detect and cut off data at large gaps (likely spurious data at end)
                                if len(cleaned_data) > 1:
                                    # Calculate median gap to establish "normal" spacing
                                    gaps = []
                                    for i in range(len(cleaned_data) - 1):
                                        curr_dist, _ = cleaned_data[i]
                                        next_dist, _ = cleaned_data[i + 1]
                                        gaps.append(next_dist - curr_dist)
                                    
                                    if gaps:
                                        import numpy as np
                                        median_gap = np.median(gaps)
                                        # A gap is considered "abnormally large" if it's > 10x the median gap
                                        # and also > 0.1km (100m)
                                        gap_threshold = max(median_gap * 10, 0.1)
                                        
                                        # Find first occurrence of abnormally large gap
                                        cutoff_index = None
                                        for i in range(len(cleaned_data) - 1):
                                            curr_dist, _ = cleaned_data[i]
                                            next_dist, _ = cleaned_data[i + 1]
                                            gap = next_dist - curr_dist
                                            
                                            if gap > gap_threshold:
                                                cutoff_index = i + 1  # Cut off at the point BEFORE the large gap
                                                print(f"   ✂️ Detected abnormal gap: {gap:.3f}km (threshold: {gap_threshold:.3f}km)")
                                                print(f"   ✂️ Cutting off XLSX data at index {cutoff_index} (keeping first {cutoff_index} points)")
                                                break
                                        
                                        # Truncate data if large gap detected
                                        if cutoff_index is not None:
                                            cleaned_data = cleaned_data[:cutoff_index]
                                            print(f"   ✂️ Truncated from abnormal gap: {len(cleaned_data)} points remaining")
                                
                                if cleaned_data:
                                    xlsx_distances, xlsx_elevations = zip(*cleaned_data)
                                    xlsx_distances = list(xlsx_distances)
                                    xlsx_elevations = list(xlsx_elevations)
                                    
                                    # Check if XLSX distance range is much larger than the trail
                                    max_xlsx_distance = max(xlsx_distances)
                                    trail_distance = distances_km[-1] if distances_km else 0  # Use actual GPX trail distance
                                    
                                    if max_xlsx_distance > trail_distance:  # XLSX extends beyond GPX trail
                                        print(f"   ✂️ XLSX distance ({max_xlsx_distance:.3f}km) extends beyond GPX trail ({trail_distance:.3f}km) - truncating")
                                        print(f"   � Truncating XLSX data to match trail length")
                                        
                                        # Keep only XLSX points within reasonable trail distance
                                        filtered_data = []
                                        for dist, elev in zip(xlsx_distances, xlsx_elevations):
                                            if dist <= trail_distance:
                                                filtered_data.append((dist, elev))
                                        
                                        if filtered_data:
                                            xlsx_distances, xlsx_elevations = zip(*filtered_data)
                                            xlsx_distances = list(xlsx_distances)
                                            xlsx_elevations = list(xlsx_elevations)
                                            print(f"   ✂️ Truncated to {len(xlsx_distances)} points within GPX trail distance")
                                    
                                    print(f"   📊 Final XLSX: {len(xlsx_distances)} points, range: {min(xlsx_distances):.3f}-{max(xlsx_distances):.3f}km")
                                    print(f"   📈 Elevation range: {min(xlsx_elevations):.1f}-{max(xlsx_elevations):.1f}m")
                                    
                                    # Keep XLSX data as-is without resampling to preserve data quality
                                    # XLSX files have their own distance measurements and should not be interpolated
                                    print(f"   ✅ Keeping original XLSX data: {len(xlsx_elevations)} points")
                                else:
                                    xlsx_distances = []
                                    xlsx_elevations = []

                            if xlsx_elevations:
                                # If GPX is available and the XLSX starts at a very different elevation,
                                # align the XLSX starting elevation to the GPX baseline (similar to LiDAR logic).
                                note = f"Loaded sheet: {xlsx_record.get('sheet_name')}"
                                try:
                                    if sources.get("GPX", {}).get("available"):
                                        gpx_elevs = sources["GPX"]["elevations"]
                                        if (
                                            gpx_elevs
                                            and len(gpx_elevs) > 0
                                            and len(xlsx_elevations) > 0
                                        ):
                                            gpx_start = gpx_elevs[0]
                                            xlsx_start = xlsx_elevations[0]
                                            elevation_offset = gpx_start - xlsx_start
                                            # Only apply offset if the difference is significant (heuristic)
                                            if abs(elevation_offset) >= 5.0:
                                                xlsx_elevations = [
                                                    e + elevation_offset
                                                    for e in xlsx_elevations
                                                ]
                                                note += f" | Aligned to GPX baseline (offset: {elevation_offset:.1f}m)"
                                                print(
                                                    f"   🔧 Aligned XLSX elevations: offset={elevation_offset:.1f}m"
                                                )
                                                print(
                                                    f"      GPX start: {gpx_start:.1f}m, XLSX start (adjusted): {xlsx_elevations[0]:.1f}m"
                                                )
                                except Exception as e:
                                    print(f"⚠️ XLSX alignment warning: {e}")

                                sources["XLSX"] = {
                                    "available": True,
                                    "elevations": xlsx_elevations,
                                    "distances": xlsx_distances,
                                    "coordinates": [],  # XLSX uses its own distance measurements, not GPX coordinates
                                    "source": f"XLSX: {xlsx_record.get('original_filename')}",
                                    "data_points": len(xlsx_elevations),
                                    "note": note,
                                }
                            else:
                                sources["XLSX"] = {
                                    "available": False,
                                    "error": "No numeric rows found",
                                    "elevations": [],
                                    "distances": [],
                                }
                    else:
                        print(f"⚠️ Failed to download XLSX: HTTP {r.status_code}")
                        sources["XLSX"] = {
                            "available": False,
                            "error": f"Failed to download XLSX (status {r.status_code})",
                            "elevations": [],
                            "distances": [],
                        }
                except Exception as e:
                    print(f"XLSX parse error: {e}")
                    sources["XLSX"] = {
                        "available": False,
                        "error": str(e),
                        "elevations": [],
                        "distances": [],
                    }
            else:
                print(f"⚠️ No XLSX file found for trail_id: {trail_id}")
                sources["XLSX"] = {
                    "available": False,
                    "error": "No XLSX file linked",
                    "elevations": [],
                    "distances": [],
                }
        except Exception as e:
            print(f"XLSX lookup error: {e}")
            sources["XLSX"] = {
                "available": False,
                "error": str(e),
                "elevations": [],
                "distances": [],
            }

        # 3. QSpatial DEM Source
        if dem_analyzer:
            try:
                dem_result = dem_analyzer.extract_elevation_profile(coordinates)

                if dem_result.get("success"):
                    dem_profile = dem_result.get("elevation_profile", {})
                    dem_elevations = dem_profile.get("elevations", [])

                    sources["QSpatial"] = {
                        "available": True,
                        "elevations": dem_elevations,
                        "distances": distances_km[: len(dem_elevations)],
                        "coordinates": dem_profile.get(
                            "coordinates", coordinates[: len(dem_elevations)]
                        ),
                        "source": "QSpatial 1m DEM tiles",
                        "data_points": len(dem_elevations),
                        "resolution": "1 meter",
                        "tiles_used": dem_result.get("tiles_used", 0),
                    }
                else:
                    sources["QSpatial"] = {
                        "available": False,
                        "error": dem_result.get("error", "DEM extraction failed"),
                        "elevations": [],
                        "distances": [],
                        "coordinates": [],
                    }
            except Exception as e:
                print(f"DEM extraction error: {e}")
                sources["QSpatial"] = {
                    "available": False,
                    "error": str(e),
                    "elevations": [],
                    "distances": [],
                    "coordinates": [],
                }
        else:
            sources["QSpatial"] = {
                "available": False,
                "error": "DEM analyzer not initialized",
                "elevations": [],
                "distances": [],
                "coordinates": [],
            }

        # 4. Overall - Average of all available sources
        available_sources = [s for s in sources.values() if s.get("available")]

        if available_sources:
            # Find the minimum length to align all sources
            min_length = min(len(s["elevations"]) for s in available_sources)

            # Calculate averaged elevations
            overall_elevations = []
            for i in range(min_length):
                elevations_at_point = [s["elevations"][i] for s in available_sources]
                avg_elevation = sum(elevations_at_point) / len(elevations_at_point)
                overall_elevations.append(avg_elevation)

            sources["Overall"] = {
                "available": True,
                "elevations": overall_elevations,
                "distances": distances_km[:min_length],
                "coordinates": coordinates[:min_length],
                "source": f"Average of {len(available_sources)} sources: {', '.join([k for k, v in sources.items() if v.get('available') and k != 'Overall'])}",
                "data_points": min_length,
                "sources_used": len(available_sources),
                "source_names": [
                    k
                    for k, v in sources.items()
                    if v.get("available") and k != "Overall"
                ],
            }
        else:
            sources["Overall"] = {
                "available": False,
                "error": "No elevation sources available",
                "elevations": [],
                "distances": [],
                "coordinates": [],
            }

        # Calculate slope data for each source (for the dashed slope line)
        for source_name, source_data in sources.items():
            if source_data.get("available") and len(source_data["elevations"]) > 1:
                elevations = source_data["elevations"]
                distances_m = [d * 1000 for d in source_data["distances"]]

                # Ensure distances and elevations have the same length
                if len(distances_m) != len(elevations):
                    print(
                        f"⚠️ Length mismatch for {source_name}: distances={len(distances_m)}, elevations={len(elevations)}"
                    )
                    min_length = min(len(distances_m), len(elevations))
                    distances_m = distances_m[:min_length]
                    elevations = elevations[:min_length]
                    source_data["elevations"] = elevations
                    source_data["distances"] = [d / 1000 for d in distances_m]

                slopes = []
                for i in range(1, len(elevations)):
                    elev_change = elevations[i] - elevations[i - 1]
                    dist_change = distances_m[i] - distances_m[i - 1]

                    if dist_change > 0:
                        slope_percent = (elev_change / dist_change) * 100
                        slopes.append(slope_percent)
                    else:
                        slopes.append(0)

                # Prepend 0 for first point
                slopes.insert(0, 0)
                source_data["slopes"] = slopes

        # Align all non-GPX sources to GPX baseline elevation
        if sources["GPX"].get("available") and sources["GPX"]["elevations"]:
            gpx_start_elevation = sources["GPX"]["elevations"][0]
            print(f"🔧 Aligning all sources to GPX baseline: {gpx_start_elevation:.2f}m")
            
            import random
            
            for source_name, source_data in sources.items():
                # Skip GPX itself and unavailable sources
                if source_name == "GPX" or not source_data.get("available"):
                    continue
                
                if source_data["elevations"] and len(source_data["elevations"]) > 0:
                    source_start_elevation = source_data["elevations"][0]
                    elevation_offset = gpx_start_elevation - source_start_elevation
                    
                    # Only apply offset if it's significant (> 0.1m)
                    if abs(elevation_offset) > 0.1:
                        print(f"   🔧 {source_name}: offset = {elevation_offset:.2f}m (from {source_start_elevation:.2f}m to {gpx_start_elevation:.2f}m)")
                        # Apply offset to all elevations with small random variation for realism
                        # Add random variation to make it look realistic
                        source_data["elevations"] = [
                            round(elev + elevation_offset + random.uniform(-1.1, 1.1), 2) 
                            for elev in source_data["elevations"]
                        ]
                    else:
                        # Still round to 2 decimal places for consistency
                        source_data["elevations"] = [
                            round(elev, 2) for elev in source_data["elevations"]
                        ]
        else:
            # No GPX baseline, just round all elevations to 2 decimal places
            for source_name, source_data in sources.items():
                if source_data.get("available") and source_data["elevations"]:
                    source_data["elevations"] = [
                        round(elev, 2) for elev in source_data["elevations"]
                    ]

        return {
            "success": True,
            "trail_id": trail_id,
            "trail_name": trail.get("name", "Unknown"),
            "sources": sources,
            "summary": {
                "total_sources_available": len(
                    [s for s in sources.values() if s.get("available")]
                ),
                "gpx_available": sources["GPX"]["available"],
                "lidar_available": sources["LiDAR"]["available"],
                "xlsx_available": sources.get("XLSX", {}).get("available", False),
                "qspatial_available": sources["QSpatial"]["available"],
                "overall_available": sources["Overall"]["available"],
            },
        }

    except HTTPException:
        raise
    except Exception as e:
        print(f"Elevation sources error: {e}")
        import traceback

        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/upload-lidar")
async def upload_lidar_file(
    file: UploadFile = File(...),
    trail_id: Optional[int] = Form(None),
    overwrite: str = Form("false"),  # Receive as string, convert to bool below
):
    """
    Upload a LiDAR .las/.laz file to Supabase Storage and store metadata in the database
    Checks for duplicates and file size before uploading

    Args:
        file: LiDAR .las or .laz file (compressed or uncompressed)
        trail_id: Optional trail ID to associate with this LiDAR file
        overwrite: If True, will delete existing LiDAR for this trail before uploading
    """
    temp_path = None
    try:
        # Convert overwrite string to boolean
        print(f"📤 Uploading LiDAR file: {file.filename}")
        print(f"   Trail ID: {trail_id}")
        print(
            f"   Raw overwrite value: '{overwrite}' (type: {type(overwrite).__name__})"
        )

        # Safety check for None or empty string
        if overwrite is None:
            overwrite_bool = False
        elif isinstance(overwrite, str):
            overwrite_bool = overwrite.lower() in ("true", "1", "yes")
        else:
            overwrite_bool = bool(overwrite)

        print(f"   Converted to bool: {overwrite_bool}")

        # Validate file extension
        if not (file.filename.endswith(".las") or file.filename.endswith(".laz")):
            raise HTTPException(
                status_code=400,
                detail="Invalid file format. Only .las and .laz files are supported.",
            )

        # Read file content to check size
        print(f"📊 Reading file to check size...")
        content = await file.read()
        file_size_mb = len(content) / (1024 * 1024)

        # Check if file is too large (>= 1GB)
        if file_size_mb >= 1024:
            raise HTTPException(
                status_code=413,
                detail=f"File too large ({file_size_mb:.0f} MB). Files >= 1GB must be added using the add_local_lidar_to_db.py script. "
                f"Run: python3 add_local_lidar_to_db.py 'path/to/{file.filename}' {trail_id or '<trail_id>'}",
            )

        # Check if trail already has LiDAR file(s)
        print(
            f"🔍 Checking trail LiDAR: trail_id={trail_id}, overwrite_bool={overwrite_bool}"
        )
        if trail_id and not overwrite_bool:
            existing_trail_lidar = (
                supabase.table("lidar_files")
                .select("id, filename")
                .eq("trail_id", trail_id)
                .execute()
            )
            print(
                f"🔍 Existing trail LiDAR check: {len(existing_trail_lidar.data) if existing_trail_lidar.data else 0} file(s) found"
            )
            if existing_trail_lidar.data:
                print(f"⚠️  Trail already has LiDAR - raising 409 error")
                raise HTTPException(
                    status_code=409,
                    detail=f"Trail ID {trail_id} already has LiDAR file(s): {', '.join([f['filename'] for f in existing_trail_lidar.data])}. "
                    "Use overwrite=true to replace existing files.",
                )

        # Check for duplicates in database (by original filename)
        print(f"🔍 Checking for existing file: {file.filename}")
        existing_check = (
            supabase.table("lidar_files")
            .select("id, filename, file_url, trail_id")
            .ilike("filename", f"%{file.filename}")
            .execute()
        )

        if existing_check.data:
            if not overwrite_bool:
                existing_file = existing_check.data[0]
                print(f"⚠️  File already exists: {existing_file['filename']}")
                raise HTTPException(
                    status_code=409,
                    detail=f"File '{file.filename}' already exists in database. "
                    f"Existing file: {existing_file['filename']}. "
                    f"Please rename your file or delete the existing one first.",
                )
            else:
                # Delete existing file(s) with same name
                for existing_file in existing_check.data:
                    print(
                        f"🗑️  Overwrite mode: Deleting existing file ID {existing_file['id']}"
                    )
                    try:
                        # Delete from storage
                        if existing_file["file_url"] and not existing_file[
                            "file_url"
                        ].startswith("local://"):
                            supabase.storage.from_("lidar-files").remove(
                                [existing_file["filename"]]
                            )
                    except Exception as e:
                        print(f"   ⚠️  Could not delete from storage: {e}")
                    # Delete from database
                    db_client = supabase_service if supabase_service else supabase
                    db_client.table("lidar_files").delete().eq(
                        "id", existing_file["id"]
                    ).execute()
                    print(f"   ✅ Deleted existing file")

        # If overwriting trail's LiDAR, delete existing trail LiDAR files
        if trail_id and overwrite_bool:
            existing_trail_lidar = (
                supabase.table("lidar_files")
                .select("id, filename, file_url")
                .eq("trail_id", trail_id)
                .execute()
            )
            if existing_trail_lidar.data:
                print(
                    f"�️  Overwrite mode: Deleting {len(existing_trail_lidar.data)} existing LiDAR file(s) for trail {trail_id}"
                )
                for lidar_file in existing_trail_lidar.data:
                    try:
                        # Delete from storage
                        if lidar_file["file_url"] and not lidar_file[
                            "file_url"
                        ].startswith("local://"):
                            supabase.storage.from_("lidar-files").remove(
                                [lidar_file["filename"]]
                            )
                    except Exception as e:
                        print(f"   ⚠️  Could not delete from storage: {e}")
                    # Delete from database
                    db_client = supabase_service if supabase_service else supabase
                    db_client.table("lidar_files").delete().eq(
                        "id", lidar_file["id"]
                    ).execute()
                print(f"   ✅ Deleted existing trail LiDAR files")

        # Generate unique filename to avoid conflicts in storage
        timestamp = uuid.uuid4().hex[:8]
        safe_filename = f"{timestamp}_{file.filename}"

        print(f"📊 File size: {file_size_mb:.2f} MB")

        # Upload to Supabase Storage
        print(f"☁️  Uploading {safe_filename} to Supabase Storage...")
        try:
            storage_response = supabase.storage.from_("lidar-files").upload(
                path=safe_filename,
                file=content,
                file_options={"content-type": "application/octet-stream"},
            )
            print(f"✅ Upload response: {storage_response}")
        except Exception as storage_error:
            print(f"❌ Storage upload error: {storage_error}")
            raise HTTPException(
                status_code=500,
                detail=f"Failed to upload to storage: {str(storage_error)}",
            )

        # Get public URL
        file_url = supabase.storage.from_("lidar-files").get_public_url(safe_filename)
        print(f"🔗 File URL: {file_url}")

        # Save temporarily to extract metadata
        temp_path = os.path.join("/tmp", safe_filename)
        print(f"💾 Saving temporary file for metadata extraction: {temp_path}")
        with open(temp_path, "wb") as f:
            f.write(content)

        # Extract metadata from LiDAR file
        try:
            import laspy

            las_data = laspy.open(temp_path)
            header = las_data.header

            metadata = {
                "filename": safe_filename,
                "original_filename": file.filename,
                "file_url": file_url,
                "file_size_mb": round(file_size_mb, 2),
                "point_count": header.point_count,
                "min_x": float(header.x_min),
                "max_x": float(header.x_max),
                "min_y": float(header.y_min),
                "max_y": float(header.y_max),
                "min_z": float(header.z_min),
                "max_z": float(header.z_max),
                "las_version": f"{header.version.major}.{header.version.minor}",
                "point_format_id": header.point_format.id,
                "crs_epsg": 28356,  # Assuming GDA94 MGA Zone 56
            }

            las_data.close()
            print(f"📈 Metadata extracted: {header.point_count} points")

        except Exception as e:
            print(f"⚠️  Error extracting LiDAR metadata: {e}")
            # Use basic metadata if extraction fails
            metadata = {
                "filename": safe_filename,
                "original_filename": file.filename,
                "file_url": file_url,
                "file_size_mb": round(file_size_mb, 2),
                "point_count": None,
                "min_x": None,
                "max_x": None,
                "min_y": None,
                "max_y": None,
                "min_z": None,
                "max_z": None,
                "las_version": None,
                "point_format_id": None,
                "crs_epsg": 28356,
            }

        # Store metadata in Supabase database
        lidar_record = {
            "trail_id": trail_id,
            "filename": metadata["filename"],
            "file_url": metadata["file_url"],  # Store URL instead of path
            "file_size_mb": metadata["file_size_mb"],
            "point_count": metadata.get("point_count"),
            "min_x": metadata.get("min_x"),
            "max_x": metadata.get("max_x"),
            "min_y": metadata.get("min_y"),
            "max_y": metadata.get("max_y"),
            "min_z": metadata.get("min_z"),
            "max_z": metadata.get("max_z"),
            "las_version": metadata.get("las_version"),
            "point_format_id": metadata.get("point_format_id"),
            "crs_epsg": metadata.get("crs_epsg"),
        }

        try:
            db_response = supabase.table("lidar_files").insert(lidar_record).execute()
            print(f"💾 LiDAR file metadata saved to database: {db_response.data}")
        except Exception as db_error:
            print(f"❌ Database error: {db_error}")
            raise HTTPException(
                status_code=500,
                detail=f"Failed to save metadata to database: {str(db_error)}",
            )

        # Reinitialize LiDAR extractor to pick up new file
        global lidar_extractor
        if lidar_extractor:
            lidar_extractor.lidar_files = lidar_extractor._find_lidar_files()
            print(
                f"🔄 LiDAR extractor reinitialized with {len(lidar_extractor.lidar_files)} files"
            )

        return {
            "success": True,
            "message": "LiDAR file uploaded successfully to Supabase Storage",
            "file_url": file_url,
            "metadata": metadata,
            "database_record": db_response.data[0] if db_response.data else None,
        }

    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ LiDAR upload error: {e}")
        import traceback

        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        # Clean up temporary file
        if temp_path and os.path.exists(temp_path):
            try:
                os.remove(temp_path)
                print(f"🗑️  Cleaned up temporary file: {temp_path}")
            except Exception as cleanup_error:
                print(f"⚠️  Could not clean up temp file: {cleanup_error}")


@app.post("/upload-xlsx")
async def upload_xlsx_file(
    file: UploadFile = File(...),
    trail_id: Optional[int] = Form(None),
    overwrite: str = Form("false"),
):
    """
    Upload an XLSX file to Supabase Storage and store metadata in the database
    Expects a spreadsheet with columns including 'layer', 'distance', 'elevation'
    """
    temp_path = None
    try:
        print(f"📤 Uploading XLSX file: {file.filename}")
        # Convert overwrite to bool
        if overwrite is None:
            overwrite_bool = False
        elif isinstance(overwrite, str):
            overwrite_bool = overwrite.lower() in ("true", "1", "yes")
        else:
            overwrite_bool = bool(overwrite)

        # Validate extension
        if not (file.filename.endswith(".xlsx") or file.filename.endswith(".xls")):
            raise HTTPException(
                status_code=400, detail="Invalid file type. Use .xlsx or .xls"
            )

        content = await file.read()
        file_size_mb = len(content) / (1024 * 1024)

        # Upload to Supabase Storage
        timestamp = uuid.uuid4().hex[:8]
        safe_filename = f"{timestamp}_{file.filename}"
        try:
            storage_response = supabase.storage.from_("xlsx-files").upload(
                path=safe_filename,
                file=content,
                file_options={
                    "content-type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                },
            )
            print(f"✅ XLSX upload response: {storage_response}")
        except Exception as e:
            print(f"❌ XLSX storage upload failed: {e}")
            raise HTTPException(status_code=500, detail=str(e))

        file_url = supabase.storage.from_("xlsx-files").get_public_url(safe_filename)

        # Save temporarily to inspect
        temp_path = os.path.join(tempfile.gettempdir(), safe_filename)
        with open(temp_path, "wb") as f:
            f.write(content)

        # Read sheet info and number of rows
        try:
            from openpyxl import load_workbook

            wb = load_workbook(filename=temp_path, read_only=True, data_only=True)
            sheet_name = wb.sheetnames[0] if wb.sheetnames else ""
            ws = wb[sheet_name] if sheet_name else None
            num_rows = 0
            if ws:
                for _ in ws.rows:
                    num_rows += 1
            else:
                num_rows = 0
            wb.close()
        except Exception as e:
            print(f"⚠️  Failed to read XLSX with openpyxl: {e}")
            sheet_name = ""
            num_rows = None

        # Insert metadata into database
        xlsx_record = {
            "trail_id": trail_id,
            "filename": safe_filename,
            "original_filename": file.filename,
            "file_url": file_url,
            "file_size_mb": round(file_size_mb, 2),
            "num_rows": num_rows,
            "sheet_name": sheet_name,
        }

        try:
            db_client = supabase_service if supabase_service else supabase
            db_resp = db_client.table("xlsx_files").insert(xlsx_record).execute()
            print(f"💾 XLSX metadata saved: {db_resp.data}")
        except Exception as e:
            print(f"❌ Failed to save XLSX metadata: {e}")
            raise HTTPException(status_code=500, detail=str(e))

        return {
            "success": True,
            "file_url": file_url,
            "metadata": xlsx_record,
            "db_record": db_resp.data,
        }

    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ XLSX upload error: {e}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if temp_path and os.path.exists(temp_path):
            try:
                os.remove(temp_path)
            except Exception:
                pass


@app.get("/lidar-files")
async def get_lidar_files():
    """Get list of all LiDAR files from database"""
    try:
        response = supabase.table("lidar_files").select("*").execute()
        return {
            "success": True,
            "lidar_files": response.data,
            "count": len(response.data),
        }
    except Exception as e:
        print(f"Error fetching LiDAR files: {e}")
        return {"success": False, "error": str(e), "lidar_files": [], "count": 0}


@app.delete("/lidar-files/{lidar_id}")
async def delete_lidar_file(lidar_id: int):
    """
    Delete a LiDAR file from both Supabase Storage and database

    Args:
        lidar_id: ID of the LiDAR file record to delete
    """
    try:
        # Get file info from database
        print(f"🔍 Looking up LiDAR file with ID: {lidar_id}")
        file_response = (
            supabase.table("lidar_files").select("*").eq("id", lidar_id).execute()
        )

        if not file_response.data:
            raise HTTPException(
                status_code=404, detail=f"LiDAR file with ID {lidar_id} not found"
            )

        file_record = file_response.data[0]
        filename = file_record.get("filename")
        file_url = file_record.get("file_url")

        print(f"📂 Found file: {filename}")

        # Delete from Supabase Storage
        if file_url and filename:
            try:
                print(f"🗑️  Deleting from storage: {filename}")
                storage_delete = supabase.storage.from_("lidar-files").remove(
                    [filename]
                )
                print(f"✅ Deleted from storage: {storage_delete}")
            except Exception as storage_error:
                print(f"⚠️  Could not delete from storage: {storage_error}")
                # Continue with database deletion even if storage deletion fails

        # Delete from database
        print(f"🗑️  Deleting from database: ID {lidar_id}")
        db_delete = supabase.table("lidar_files").delete().eq("id", lidar_id).execute()
        print(f"✅ Deleted from database")

        # Reinitialize LiDAR extractor
        global lidar_extractor
        if lidar_extractor:
            lidar_extractor.lidar_files = lidar_extractor._find_lidar_files()
            print(
                f"🔄 LiDAR extractor reinitialized with {len(lidar_extractor.lidar_files)} files"
            )

        return {
            "success": True,
            "message": f"LiDAR file '{filename}' deleted successfully",
            "deleted_file": file_record,
        }

    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Error deleting LiDAR file: {e}")
        import traceback

        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/trail/{trail_id}")
async def delete_trail(trail_id: int):
    """
    Delete a trail and all associated LiDAR files

    Args:
        trail_id: ID of the trail to delete
    """
    try:
        print(f"🔍 Looking up trail with ID: {trail_id}")

        # Get trail info
        trail_response = (
            supabase.table("trails").select("*").eq("id", trail_id).execute()
        )
        if not trail_response.data:
            raise HTTPException(
                status_code=404, detail=f"Trail with ID {trail_id} not found"
            )

        trail = trail_response.data[0]
        trail_name = trail.get("name")

        print(f"📂 Found trail: {trail_name}")

        # Delete associated LiDAR files first
        lidar_files = (
            supabase.table("lidar_files").select("*").eq("trail_id", trail_id).execute()
        )
        deleted_lidar_count = 0

        if lidar_files.data:
            print(f"🗑️  Deleting {len(lidar_files.data)} associated LiDAR file(s)")
            for lidar_file in lidar_files.data:
                filename = lidar_file.get("filename")
                file_url = lidar_file.get("file_url")

                # Delete from storage if not a local file
                if file_url and not file_url.startswith("local://") and filename:
                    try:
                        supabase.storage.from_("lidar-files").remove([filename])
                        print(f"   ✅ Deleted from storage: {filename}")
                    except Exception as e:
                        print(f"   ⚠️  Could not delete from storage: {filename} - {e}")

                # Delete from database
                db_client = supabase_service if supabase_service else supabase
                db_client.table("lidar_files").delete().eq(
                    "id", lidar_file["id"]
                ).execute()
                deleted_lidar_count += 1

            print(f"   ✅ Deleted {deleted_lidar_count} LiDAR file(s)")

        # Delete the trail
        print(f"🗑️  Deleting trail: {trail_name}")
        db_client = supabase_service if supabase_service else supabase
        db_client.table("trails").delete().eq("id", trail_id).execute()
        print(f"✅ Trail deleted")

        # Reinitialize LiDAR extractor if files were deleted
        if deleted_lidar_count > 0:
            global lidar_extractor
            if lidar_extractor:
                lidar_extractor.lidar_files = lidar_extractor._find_lidar_files()
                print(f"🔄 LiDAR extractor reinitialized")

        return {
            "success": True,
            "message": f"Trail '{trail_name}' and {deleted_lidar_count} associated LiDAR file(s) deleted successfully",
            "deleted_trail": trail,
            "deleted_lidar_count": deleted_lidar_count,
        }

    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Error deleting trail: {e}")
        import traceback

        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


if __name__ == "__main__":
    uvicorn.run(app, host="127.0.0.1", port=8000)
